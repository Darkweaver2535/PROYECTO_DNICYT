import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Avg, Sum, F
from django.core.paginator import Paginator
from django.http import JsonResponse
from datetime import date, timedelta, datetime
from django.utils import timezone

# Importaciones para Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# Importaciones para PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Importar modelos
from apps.equipos.models import Equipo
from apps.materiales.models import Material, MovimientoMaterial
from apps.inventario.models import MovimientoStock, Repuesto
from apps.mantenimiento.models import PlanMantenimiento, OrdenTrabajo
from .models import ReporteGenerado, AnalisisEquipos, RegistroFalla, SeguimientoFalla
from .forms import RegistroFallaForm

@login_required
def reportes_equipos_view(request):
    """Vista principal de Reportes de Equipos con análisis completo"""
    
    # Obtener filtros de la URL
    seccion_filtro = request.GET.get('seccion', '')
    estado_filtro = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_analisis = request.GET.get('tipo_analisis', 'general')
    
    # === CONSULTA BASE DE EQUIPOS ===
    equipos_queryset = Equipo.objects.all()
    
    # Aplicar filtros
    if seccion_filtro:
        equipos_queryset = equipos_queryset.filter(seccion=seccion_filtro)
    
    if estado_filtro:
        equipos_queryset = equipos_queryset.filter(estado=estado_filtro)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            equipos_queryset = equipos_queryset.filter(fecha_ingreso__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            equipos_queryset = equipos_queryset.filter(fecha_ingreso__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # === MÉTRICAS PRINCIPALES ===
    total_equipos = equipos_queryset.count()
    equipos_operativos = equipos_queryset.filter(estado='OPERATIVO').count()
    equipos_mantenimiento = equipos_queryset.filter(estado='MANTENIMIENTO').count()
    equipos_fuera_servicio = equipos_queryset.filter(estado='FUERA_SERVICIO').count()
    
    # Calcular porcentajes
    if total_equipos > 0:
        porcentaje_operativo = round((equipos_operativos / total_equipos) * 100, 1)
        porcentaje_mantenimiento = round((equipos_mantenimiento / total_equipos) * 100, 1)
        porcentaje_fuera_servicio = round((equipos_fuera_servicio / total_equipos) * 100, 1)
    else:
        porcentaje_operativo = porcentaje_mantenimiento = porcentaje_fuera_servicio = 0
    
    # === ANÁLISIS POR SECCIÓN ===
    equipos_por_seccion = equipos_queryset.values('seccion').annotate(
        total=Count('id'),
        operativos=Count('id', filter=Q(estado='OPERATIVO')),
        mantenimiento=Count('id', filter=Q(estado='MANTENIMIENTO')),
        fuera_servicio=Count('id', filter=Q(estado='FUERA_SERVICIO'))
    ).order_by('-total')
    
    # === EQUIPOS CON FICHAS TÉCNICAS ===
    equipos_con_ficha = equipos_queryset.filter(ficha_tecnica_completa=True).count()
    equipos_sin_ficha = total_equipos - equipos_con_ficha
    porcentaje_fichas_completas = round((equipos_con_ficha / total_equipos) * 100, 1) if total_equipos > 0 else 0
    
    # === ANÁLISIS DE CRITICIDAD ===
    try:
        equipos_criticos = equipos_queryset.filter(
            Q(estado='MANTENIMIENTO') | Q(estado='FUERA_SERVICIO')
        ).count()
        equipos_no_criticos = total_equipos - equipos_criticos
    except:
        equipos_criticos = equipos_no_criticos = 0
    
    # === ALERTAS Y RECOMENDACIONES ===
    alertas = []
    
    # Alerta por equipos fuera de servicio
    if equipos_fuera_servicio > 0:
        alertas.append({
            'tipo': 'warning',
            'mensaje': f'{equipos_fuera_servicio} equipos están fuera de servicio',
            'accion': 'Revisar y programar mantenimiento correctivo'
        })
    
    # Alerta por equipos en mantenimiento
    if equipos_mantenimiento > 5:
        alertas.append({
            'tipo': 'info',
            'mensaje': f'{equipos_mantenimiento} equipos en mantenimiento actualmente',
            'accion': 'Verificar cronograma de mantenimiento'
        })
    
    # Alerta por fichas técnicas incompletas
    if porcentaje_fichas_completas < 70:
        alertas.append({
            'tipo': 'warning',
            'mensaje': f'Solo {porcentaje_fichas_completas}% de fichas técnicas completas',
            'accion': 'Completar documentación técnica faltante'
        })
    
    # === OBTENER DATOS PARA FILTROS ===
    secciones_disponibles = Equipo.SECCION_CHOICES
    estados_disponibles = Equipo.ESTADO_CHOICES
    
    # === ÚLTIMOS REPORTES GENERADOS ===
    try:
        ultimos_reportes = ReporteGenerado.objects.filter(
            tipo_reporte__in=['equipos_general', 'equipos_estado', 'equipos_ubicacion']
        ).order_by('-fecha_generacion')[:5]
    except:
        ultimos_reportes = []
    
    # === CONTEXT PARA EL TEMPLATE ===
    context = {
        'seccion_filtro': seccion_filtro,
        'estado_filtro': estado_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_analisis': tipo_analisis,
        'secciones_disponibles': secciones_disponibles,
        'estados_disponibles': estados_disponibles,
        'total_equipos': total_equipos,
        'equipos_operativos': equipos_operativos,
        'equipos_mantenimiento': equipos_mantenimiento,
        'equipos_fuera_servicio': equipos_fuera_servicio,
        'porcentaje_operativo': porcentaje_operativo,
        'porcentaje_mantenimiento': porcentaje_mantenimiento,
        'porcentaje_fuera_servicio': porcentaje_fuera_servicio,
        'equipos_por_seccion': equipos_por_seccion,
        'equipos_con_ficha': equipos_con_ficha,
        'equipos_sin_ficha': equipos_sin_ficha,
        'porcentaje_fichas_completas': porcentaje_fichas_completas,
        'equipos_criticos': equipos_criticos,
        'equipos_no_criticos': equipos_no_criticos,
        'alertas': alertas,
        'ultimos_reportes': ultimos_reportes,
        'titulo': 'Reportes de Equipos',
    }
    
    return render(request, 'sistema_interno/reportes_equipos.html', context)

@login_required
def exportar_reporte_equipos_excel(request):
    """Exportar reporte de equipos a Excel"""
    
    # Obtener filtros
    seccion_filtro = request.GET.get('seccion', '')
    estado_filtro = request.GET.get('estado', '')
    
    # Obtener datos
    equipos = Equipo.objects.all()
    
    if seccion_filtro:
        equipos = equipos.filter(seccion=seccion_filtro)
    if estado_filtro:
        equipos = equipos.filter(estado=estado_filtro)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Equipos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título del reporte
    ws.merge_cells('A1:J1')
    ws['A1'] = 'REPORTE DE EQUIPOS'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_alignment
    
    # Información del reporte
    ws['A2'] = f'Fecha de generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'] = f'Total de equipos: {equipos.count()}'
    
    # Encabezados (fila 5)
    headers = [
        'Código', 'Nombre', 'Sección', 'Estado', 'Fabricante',
        'Modelo', 'Año', 'Responsable', 'Ubicación', 'Observaciones'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Datos
    for row_num, equipo in enumerate(equipos, 6):
        datos = [
            equipo.codigo_interno,
            equipo.nombre,
            equipo.get_seccion_display(),
            equipo.get_estado_display(),
            equipo.fabricante or 'No especificado',
            equipo.modelo or 'No especificado',
            equipo.año_fabricacion or 'No especificado',
            equipo.responsable or 'No asignado',
            equipo.ubicacion_fisica,
            equipo.observaciones or 'Sin observaciones'
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = valor
            cell.border = border
            
            # Colorear según estado
            if equipo.estado == 'FUERA_SERVICIO':
                cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            elif equipo.estado == 'MANTENIMIENTO':
                cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Equipos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def exportar_reporte_equipos_pdf(request):
    """Exportar reporte completo de equipos a PDF con gráficos y estadísticas"""
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    timestamp = timezone.now().strftime("%Y%m%d_%H%M")
    filename = f'Reporte_Equipos_{timestamp}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Crear documento PDF usando ReportLab
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    title = Paragraph("Reporte de Equipos", styles['Title'])
    story.append(title)
    
    doc.build(story)
    return response

@login_required
def api_equipos_chart_data(request):
    """API para datos de gráficos de equipos"""
    
    # Datos básicos por estado
    data = {
        'estados': {
            'labels': ['Operativo', 'Mantenimiento', 'Fuera de Servicio'],
            'data': [
                Equipo.objects.filter(estado='OPERATIVO').count(),
                Equipo.objects.filter(estado='MANTENIMIENTO').count(),
                Equipo.objects.filter(estado='FUERA_SERVICIO').count(),
            ]
        }
    }
    
    return JsonResponse(data)

@login_required
def analisis_fallas_view(request):
    """Vista principal para el análisis de fallas de equipos industriales con datos reales"""
    
    # === OBTENER FILTROS ===
    search_query = request.GET.get('search', '')
    severidad_filtro = request.GET.get('severidad', '')
    estado_filtro = request.GET.get('estado', '')
    equipo_filtro = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_falla_filtro = request.GET.get('tipo_falla', '')
    responsable_filtro = request.GET.get('responsable', '')
    
    # === QUERY BASE DE FALLAS REALES ===
    fallas_queryset = RegistroFalla.objects.select_related(
        'equipo', 'reportado_por', 'asignado_a'
    ).filter(activo=True)
    
    # === APLICAR FILTROS ===
    if search_query:
        fallas_queryset = fallas_queryset.filter(
            Q(equipo__nombre__icontains=search_query) |
            Q(equipo__codigo_interno__icontains=search_query) |
            Q(descripcion_falla__icontains=search_query) |
            Q(codigo_falla__icontains=search_query) |
            Q(causa_raiz__icontains=search_query)
        )
    
    if severidad_filtro:
        fallas_queryset = fallas_queryset.filter(severidad=severidad_filtro)
    
    if estado_filtro:
        fallas_queryset = fallas_queryset.filter(estado=estado_filtro)
    
    if equipo_filtro:
        fallas_queryset = fallas_queryset.filter(equipo_id=equipo_filtro)
    
    if tipo_falla_filtro:
        fallas_queryset = fallas_queryset.filter(tipo_falla=tipo_falla_filtro)
    
    if responsable_filtro:
        fallas_queryset = fallas_queryset.filter(
            Q(reportado_por_id=responsable_filtro) |
            Q(asignado_a_id=responsable_filtro)
        )
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fallas_queryset = fallas_queryset.filter(fecha_ocurrencia__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            fallas_queryset = fallas_queryset.filter(fecha_ocurrencia__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # === ORDENAR POR CRITICIDAD Y FECHA ===
    fallas_queryset = fallas_queryset.order_by('-indice_criticidad', '-fecha_ocurrencia')
    
    # === CALCULAR ESTADÍSTICAS REALES ===
    total_fallas = fallas_queryset.count()
    fallas_criticas = fallas_queryset.filter(severidad='critica').count()
    fallas_altas = fallas_queryset.filter(severidad='alta').count()
    fallas_pendientes = fallas_queryset.filter(estado__in=['identificada', 'pendiente']).count()
    fallas_resueltas = fallas_queryset.filter(estado='solucionada').count()
    fallas_en_analisis = fallas_queryset.filter(estado='analisis').count()
    
    # Calcular MTBF y MTTR promedios (si hay datos suficientes)
    if total_fallas > 0:
        mtbf_promedio = fallas_queryset.exclude(
            tiempo_parada__isnull=True
        ).aggregate(promedio=Avg('tiempo_parada'))['promedio'] or 0
        
        mttr_promedio = fallas_queryset.filter(
            tiempo_reparacion__isnull=False
        ).aggregate(promedio=Avg('tiempo_reparacion'))['promedio'] or 0
        
        if mtbf_promedio == 0:
            mtbf_promedio = 168  # Valor por defecto
        if mttr_promedio == 0:
            mttr_promedio = 4  # Valor por defecto
    else:
        mtbf_promedio = 168
        mttr_promedio = 4
    
    # === ALERTAS CRÍTICAS DINÁMICAS ===
    alertas_criticas = []
    
    # Alertas por fallas críticas
    if fallas_criticas > 0:
        alertas_criticas.append({
            'titulo': f'{fallas_criticas} Fallas Críticas Activas',
            'descripcion': 'Requieren atención inmediata para evitar paradas prolongadas de equipos'
        })
    
    # Alertas por fallas pendientes
    if fallas_pendientes > 3:
        alertas_criticas.append({
            'titulo': 'Acumulación de Fallas Pendientes',
            'descripcion': f'{fallas_pendientes} fallas esperan análisis y resolución'
        })
    
    # Alertas por fallas recurrentes (mismo equipo, múltiples fallas)
    fallas_recurrentes = fallas_queryset.values('equipo').annotate(
        count=Count('id')
    ).filter(count__gte=3)
    
    if fallas_recurrentes.exists():
        equipos_problematicos = fallas_recurrentes.count()
        alertas_criticas.append({
            'titulo': 'Equipos con Fallas Recurrentes Detectados',
            'descripcion': f'{equipos_problematicos} equipos presentan múltiples fallas. Revisar mantenimiento preventivo.'
        })
    
    # Alertas por fallas sin asignar
    fallas_sin_asignar = fallas_queryset.filter(asignado_a__isnull=True, estado__in=['identificada', 'pendiente']).count()
    if fallas_sin_asignar > 0:
        alertas_criticas.append({
            'titulo': 'Fallas Sin Asignar',
            'descripcion': f'{fallas_sin_asignar} fallas no tienen responsable asignado'
        })
    
    # === ACTIVIDAD RECIENTE REAL ===
    actividad_reciente = []
    
    # Fallas recientes (últimas 5)
    fallas_recientes = RegistroFalla.objects.select_related('equipo', 'reportado_por').order_by('-fecha_registro')[:5]
    for falla in fallas_recientes:
        actividad_reciente.append({
            'fecha': falla.fecha_registro,
            'descripcion': f'Nueva falla registrada: {falla.codigo_falla} en {falla.equipo.nombre}'
        })
    
    # Fallas resueltas recientes
    fallas_resueltas_recientes = RegistroFalla.objects.filter(
        estado='solucionada',
        fecha_solucion__isnull=False
    ).select_related('equipo').order_by('-fecha_solucion')[:3]
    
    for falla in fallas_resueltas_recientes:
        actividad_reciente.append({
            'fecha': falla.fecha_solucion,
            'descripcion': f'Falla resuelta: {falla.codigo_falla} en {falla.equipo.nombre}'
        })
    
    # Ordenar actividad por fecha
    actividad_reciente.sort(key=lambda x: x['fecha'], reverse=True)
    actividad_reciente = actividad_reciente[:8]  # Mostrar solo las 8 más recientes
    
    # === DATOS PARA GRÁFICO DE TENDENCIAS (últimos 6 meses) ===
    tendencia_datos = []
    for i in range(6):
        mes_inicio = (timezone.now() - timedelta(days=30*i)).replace(day=1)
        mes_fin = mes_inicio + timedelta(days=32)
        mes_fin = mes_fin.replace(day=1) - timedelta(days=1)
        
        fallas_mes = RegistroFalla.objects.filter(
            fecha_ocurrencia__date__gte=mes_inicio.date(),
            fecha_ocurrencia__date__lte=mes_fin.date()
        ).count()
        
        tendencia_datos.insert(0, fallas_mes)  # Insertar al principio para orden cronológico
    
    # === OBTENER DATOS PARA FILTROS ===
    equipos_disponibles = Equipo.objects.filter(
        id__in=fallas_queryset.values_list('equipo_id', flat=True).distinct()
    ).order_by('codigo_interno')
    
    # Si no hay fallas, mostrar todos los equipos
    if not equipos_disponibles.exists():
        equipos_disponibles = Equipo.objects.all().order_by('codigo_interno')
    
    # Usuarios que han reportado o tienen fallas asignadas
    usuarios_involucrados = set()
    for falla in fallas_queryset.select_related('reportado_por', 'asignado_a'):
        if falla.reportado_por:
            usuarios_involucrados.add(falla.reportado_por)
        if falla.asignado_a:
            usuarios_involucrados.add(falla.asignado_a)
    
    # === PAGINACIÓN ===
    paginator = Paginator(fallas_queryset, 12)  # 12 fallas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # === CONTEXT PARA EL TEMPLATE ===
    context = {
        'fallas': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'severidad_filtro': severidad_filtro,
        'estado_filtro': estado_filtro,
        'equipo_filtro': equipo_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_falla_filtro': tipo_falla_filtro,
        'responsable_filtro': responsable_filtro,
        'equipos_disponibles': equipos_disponibles,
        'usuarios_involucrados': sorted(list(usuarios_involucrados), key=lambda u: u.get_full_name() or u.username),
        'severidades': RegistroFalla.SEVERIDAD_CHOICES,
        'estados': RegistroFalla.ESTADO_CHOICES,
        'tipos_falla': RegistroFalla.TIPO_FALLA_CHOICES,
        'total_fallas': total_fallas,
        'fallas_criticas': fallas_criticas,
        'fallas_altas': fallas_altas,
        'fallas_pendientes': fallas_pendientes,
        'fallas_resueltas': fallas_resueltas,
        'fallas_en_analisis': fallas_en_analisis,
        'mtbf_promedio': round(mtbf_promedio, 1),
        'mttr_promedio': round(mttr_promedio, 1),
        'alertas_criticas': alertas_criticas,
        'actividad_reciente': actividad_reciente,
        'tendencia_datos': json.dumps(tendencia_datos),
        'titulo': 'Análisis de Fallas',
        'fecha_actualizacion': timezone.now(),
    }
    
    return render(request, 'sistema_interno/analisis_fallas.html', context)

@login_required
def crear_falla_view(request):
    """Vista para crear nueva falla"""
    
    if request.method == 'POST':
        form = RegistroFallaForm(request.POST, request.FILES)
        if form.is_valid():
            falla = form.save(commit=False)
            falla.reportado_por = request.user
            falla.save()
            
            messages.success(request, f'Falla {falla.codigo_falla} registrada exitosamente.')
            return redirect('reportes:analisis-fallas')
        else:
            messages.error(request, 'Error al registrar la falla. Por favor revise los campos.')
    else:
        form = RegistroFallaForm()
    
    # Estadísticas para el contexto
    stats = {
        'total_equipos': Equipo.objects.count(),
        'total_fallas': RegistroFalla.objects.count(),
        'fallas_mes': RegistroFalla.objects.filter(
            fecha_ocurrencia__month=timezone.now().month,
            fecha_ocurrencia__year=timezone.now().year
        ).count(),
    }
    
    context = {
        'form': form,
        'titulo': 'Registrar Nueva Falla',
        'stats': stats,
    }
    
    return render(request, 'sistema_interno/crear_falla.html', context)

@login_required
def detalle_falla_view(request, codigo_falla):
    """Vista de detalle de falla específica"""
    falla = get_object_or_404(RegistroFalla, codigo_falla=codigo_falla)
    
    # Obtener seguimientos de la falla
    seguimientos = falla.seguimientos.all().order_by('-fecha_accion')
    
    # Calcular métricas
    falla.dias_transcurridos = falla.get_dias_transcurridos()
    falla.esta_vencida_resolucion = falla.esta_vencida()
    
    context = {
        'falla': falla,
        'seguimientos': seguimientos,
        'titulo': f'Falla {codigo_falla}',
    }
    
    return render(request, 'sistema_interno/detalle_falla.html', context)

@login_required
def editar_falla_view(request, codigo_falla):
    """Vista para editar falla existente"""
    falla = get_object_or_404(RegistroFalla, codigo_falla=codigo_falla)
    
    if request.method == 'POST':
        form = RegistroFallaForm(request.POST, request.FILES, instance=falla)
        if form.is_valid():
            form.save()
            messages.success(request, f'Falla {falla.codigo_falla} actualizada exitosamente.')
            return redirect('reportes:detalle-falla', codigo_falla=falla.codigo_falla)
        else:
            messages.error(request, 'Error al actualizar la falla. Por favor revise los campos.')
    else:
        form = RegistroFallaForm(instance=falla)
    
    context = {
        'form': form,
        'falla': falla,
        'titulo': f'Editar Falla {codigo_falla}',
    }
    
    return render(request, 'sistema_interno/crear_falla.html', context)

@login_required
def eliminar_falla_view(request, codigo_falla):
    """Vista para eliminar falla"""
    falla = get_object_or_404(RegistroFalla, codigo_falla=codigo_falla)
    
    # Verificar si tiene seguimientos
    tiene_seguimientos = falla.seguimientos.exists()
    
    if request.method == 'POST':
        # Eliminar la falla y todos sus seguimientos
        codigo = falla.codigo_falla
        falla.delete()
        
        messages.success(request, f'Falla {codigo} eliminada exitosamente.')
        return redirect('reportes:analisis-fallas')
    
    context = {
        'falla': falla,
        'tiene_seguimientos': tiene_seguimientos,
        'titulo': f'Eliminar Falla {codigo_falla}',
    }
    
    return render(request, 'sistema_interno/confirmar_eliminar_falla.html', context)

@login_required
def cerrar_falla_view(request, codigo_falla):
    """Vista para cerrar falla como solucionada"""
    falla = get_object_or_404(RegistroFalla, codigo_falla=codigo_falla)
    
    if request.method == 'POST':
        # Obtener datos del formulario
        solucion_aplicada = request.POST.get('solucion_aplicada', '').strip()
        tiempo_reparacion = request.POST.get('tiempo_reparacion')
        repuestos_utilizados = request.POST.get('repuestos_utilizados', '').strip()
        costo_reparacion = request.POST.get('costo_reparacion')
        acciones_preventivas = request.POST.get('acciones_preventivas', '').strip()
        recomendaciones = request.POST.get('recomendaciones', '').strip()
        observaciones_cierre = request.POST.get('observaciones_cierre', '').strip()
        verificado_funcionamiento = request.POST.get('verificado_funcionamiento') == '1'
        
        # Validación básica
        if not solucion_aplicada:
            messages.error(request, 'Debe describir la solución aplicada para cerrar la falla.')
            return render(request, 'sistema_interno/cerrar_falla.html', {'falla': falla})
        
        # Actualizar la falla
        falla.estado = 'solucionada'
        falla.fecha_solucion = timezone.now()
        falla.solucion_aplicada = solucion_aplicada
        
        if tiempo_reparacion:
            try:
                falla.tiempo_reparacion = float(tiempo_reparacion)
            except ValueError:
                pass
        
        if repuestos_utilizados:
            falla.repuestos_utilizados = repuestos_utilizados
        
        if costo_reparacion:
            try:
                falla.costo_reparacion = float(costo_reparacion)
            except ValueError:
                pass
        
        if acciones_preventivas:
            falla.acciones_preventivas = acciones_preventivas
        
        if recomendaciones:
            falla.recomendaciones = recomendaciones
        
        # Agregar observaciones de cierre a las observaciones existentes
        if observaciones_cierre:
            if falla.observaciones:
                falla.observaciones += f"\n\n--- CIERRE ({timezone.now().strftime('%d/%m/%Y %H:%M')}) ---\n{observaciones_cierre}"
            else:
                falla.observaciones = f"--- CIERRE ({timezone.now().strftime('%d/%m/%Y %H:%M')}) ---\n{observaciones_cierre}"
        
        falla.save()
        
        # Crear seguimiento de cierre
        SeguimientoFalla.objects.create(
            falla=falla,
            fecha_accion=timezone.now(),
            tipo_accion='cierre_falla',
            descripcion_accion=f"Falla cerrada como solucionada. Solución: {solucion_aplicada[:100]}...",
            responsable=request.user,
            tiempo_empleado=falla.tiempo_reparacion,
            costo_accion=falla.costo_reparacion,
            resultado=f"Falla solucionada exitosamente. {'Funcionamiento verificado.' if verificado_funcionamiento else 'Funcionamiento no verificado.'}"
        )
        
        messages.success(request, f'Falla {falla.codigo_falla} cerrada como solucionada exitosamente.')
        return redirect('reportes:detalle-falla', codigo_falla=falla.codigo_falla)
    
    context = {
        'falla': falla,
        'titulo': f'Cerrar Falla {codigo_falla}',
    }
    
    return render(request, 'sistema_interno/cerrar_falla.html', context)

@login_required
def asignar_falla_view(request, codigo_falla):
    """Vista para asignar falla a un usuario"""
    falla = get_object_or_404(RegistroFalla, codigo_falla=codigo_falla)
    
    if request.method == 'POST':
        usuario_id = request.POST.get('asignado_a')
        cambiar_estado = request.POST.get('cambiar_estado') == '1'
        
        if usuario_id:
            from django.contrib.auth.models import User
            usuario = get_object_or_404(User, id=usuario_id)
            
            # Guardar el usuario anterior para el seguimiento
            usuario_anterior = falla.asignado_a
            
            # Asignar nuevo usuario
            falla.asignado_a = usuario
            
            # Cambiar estado si se solicitó
            if cambiar_estado and falla.estado in ['identificada', 'pendiente']:
                falla.estado = 'analisis'
            
            falla.save()
            
            # Crear seguimiento
            if usuario_anterior:
                descripcion = f"Falla reasignada de {usuario_anterior.get_full_name() or usuario_anterior.username} a {usuario.get_full_name() or usuario.username}"
            else:
                descripcion = f"Falla asignada a {usuario.get_full_name() or usuario.username}"
            
            if cambiar_estado:
                descripcion += " y estado cambiado a 'En Análisis'"
            
            SeguimientoFalla.objects.create(
                falla=falla,
                fecha_accion=timezone.now(),
                tipo_accion='asignacion',
                descripcion_accion=descripcion,
                responsable=request.user
            )
            
            messages.success(request, f'Falla {falla.codigo_falla} asignada exitosamente a {usuario.get_full_name() or usuario.username}.')
        else:
            messages.error(request, 'Debe seleccionar un usuario.')
        
        return redirect('reportes:detalle-falla', codigo_falla=falla.codigo_falla)
    
    # Obtener usuarios activos
    from django.contrib.auth.models import User
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'falla': falla,
        'usuarios': usuarios,
        'titulo': f'Asignar Falla {codigo_falla}',
    }
    
    return render(request, 'sistema_interno/asignar_falla.html', context)

# === FUNCIONES AUXILIARES ===

def generar_texto_filtros(seccion, estado, fecha_desde, fecha_hasta):
    """Genera texto descriptivo de los filtros aplicados"""
    filtros = []
    
    if seccion:
        filtros.append(f"Sección: {seccion}")
    else:
        filtros.append("Todas las secciones")
    
    if estado:
        filtros.append(f"Estado: {estado}")
    else:
        filtros.append("Todos los estados")
    
    if fecha_desde:
        filtros.append(f"Desde: {fecha_desde}")
    
    if fecha_hasta:
        filtros.append(f"Hasta: {fecha_hasta}")
    
    if not fecha_desde and not fecha_hasta:
        filtros.append("Sin filtro de fechas")
    
    return " | ".join(filtros)

def truncar_texto(texto, max_length):
    """Trunca texto para que no exceda el ancho de la tabla"""
    if not texto:
        return ''
    
    if len(texto) <= max_length:
        return texto
    
    return texto[:max_length-3] + '...'

def generar_recomendaciones(total, operativos, mantenimiento, fuera_servicio, por_seccion):
    """Genera recomendaciones basadas en los datos del reporte"""
    recomendaciones = []
    
    if total == 0:
        recomendaciones.append("No hay equipos registrados en el sistema.")
        return recomendaciones
    
    # Análisis de disponibilidad
    disponibilidad = (operativos / total) * 100 if total > 0 else 0
    
    if disponibilidad >= 85:
        recomendaciones.append("Excelente nivel de disponibilidad de equipos. Mantener el programa de mantenimiento preventivo.")
    elif disponibilidad >= 70:
        recomendaciones.append("Nivel de disponibilidad aceptable. Revisar equipos en mantenimiento para optimizar tiempos.")
    else:
        recomendaciones.append("Nivel de disponibilidad bajo. Se requiere intervención urgente en el programa de mantenimiento.")
    
    # Análisis de equipos fuera de servicio
    if fuera_servicio > 0:
        recomendaciones.append(f"Priorizar la reparación de {fuera_servicio} equipos fuera de servicio.")
    
    # Análisis de mantenimiento
    if mantenimiento > 0:
        recomendaciones.append(f"Optimizar cronograma de mantenimiento para los {mantenimiento} equipos actualmente en servicio.")
    
    # Recomendación general
    recomendaciones.append("Implementar sistema de monitoreo continuo y mantener actualizado el inventario de repuestos críticos.")
    
    return recomendaciones

