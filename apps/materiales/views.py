from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, F
from django.core.paginator import Paginator
from django.http import JsonResponse
from .models import Material, CategoriaMaterial, MovimientoMaterial
from .forms import MaterialForm, HerramientaForm, MovimientoMaterialForm
from apps.inventario.models import Proveedor
from datetime import date, timedelta
from decimal import Decimal

@login_required
def materiales_view(request):
    """Vista principal de materiales"""
    
    # Filtros
    search_query = request.GET.get('search', '')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    criticidad_filtro = request.GET.get('criticidad', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    tipo_filtro = request.GET.get('tipo', '')
    
    # Query base
    materiales = Material.objects.select_related('categoria', 'proveedor_principal').all()
    
    # Aplicar filtros
    if search_query:
        materiales = materiales.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query)
        )
    
    if categoria_filtro:
        materiales = materiales.filter(categoria_id=categoria_filtro)
    
    if estado_filtro:
        materiales = materiales.filter(estado=estado_filtro)
    
    if criticidad_filtro:
        materiales = materiales.filter(criticidad=criticidad_filtro)
    
    if proveedor_filtro:
        materiales = materiales.filter(proveedor_principal_id=proveedor_filtro)
    
    if tipo_filtro:
        materiales = materiales.filter(tipo=tipo_filtro)
    
    # Calcular métricas para cada material
    for material in materiales:
        material.valor_total = material.valor_stock_actual()
        material.necesita_reorden = material.necesita_reposicion()
        material.dias_vencimiento = material.dias_hasta_vencimiento()
    
    # Estadísticas
    total_materiales = materiales.count()
    materiales_disponibles = materiales.filter(estado='disponible').count()
    materiales_agotados = materiales.filter(estado='agotado').count()
    materiales_bajo_stock = materiales.filter(stock_actual__lte=F('stock_minimo')).count()
    materiales_criticos = materiales.filter(criticidad='critica').count()
    
    # Valor total del inventario
    valor_total_inventario = sum(m.valor_total for m in materiales if m.valor_total)
    
    # Materiales próximos a vencer (30 días)
    proximos_vencer = materiales.filter(
        fecha_vencimiento__lte=date.today() + timedelta(days=30),
        fecha_vencimiento__gt=date.today()
    ).count()
    
    # Alertas
    alertas = []
    
    if materiales_agotados > 0:
        alertas.append({
            'tipo': 'danger',
            'icono': 'bi-exclamation-triangle-fill',
            'titulo': f'{materiales_agotados} Material(es) Agotado(s)',
            'descripcion': 'Materiales sin stock que requieren reposición inmediata.'
        })
    
    if materiales_bajo_stock > 0:
        alertas.append({
            'tipo': 'warning',
            'icono': 'bi-arrow-down-circle',
            'titulo': f'{materiales_bajo_stock} Material(es) con Stock Bajo',
            'descripcion': 'Materiales por debajo del nivel mínimo de stock.'
        })
    
    if proximos_vencer > 0:
        alertas.append({
            'tipo': 'info',
            'icono': 'bi-calendar-event',
            'titulo': f'{proximos_vencer} Material(es) Próximo(s) a Vencer',
            'descripcion': 'Materiales que vencen en los próximos 30 días.'
        })
    
    # Obtener datos para filtros
    categorias = CategoriaMaterial.objects.filter(activo=True).order_by('nombre')
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    # Paginación
    paginator = Paginator(materiales, 12)  # 12 materiales por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'materiales': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'criticidad_filtro': criticidad_filtro,
        'proveedor_filtro': proveedor_filtro,
        'tipo_filtro': tipo_filtro,
        'categorias': categorias,
        'proveedores': proveedores,
        'estados': Material.ESTADO_CHOICES,
        'criticidades': Material.CRITICIDAD_CHOICES,
        'tipos': Material.TIPO_CHOICES,
        'total_materiales': total_materiales,
        'materiales_disponibles': materiales_disponibles,
        'materiales_agotados': materiales_agotados,
        'materiales_bajo_stock': materiales_bajo_stock,
        'materiales_criticos': materiales_criticos,
        'proximos_vencer': proximos_vencer,
        'valor_total_inventario': valor_total_inventario,
        'alertas': alertas,
        'titulo': 'Materiales',
    }
    
    return render(request, 'sistema_interno/materiales.html', context)

@login_required
def herramientas_view(request):
    """Vista principal de herramientas"""
    
    # Filtros
    search_query = request.GET.get('search', '')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    criticidad_filtro = request.GET.get('criticidad', '')
    calibracion_filtro = request.GET.get('calibracion', '')
    
    # Query base - solo herramientas
    herramientas = Material.objects.select_related('categoria', 'proveedor_principal').filter(
        tipo__in=[
            'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
            'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
            'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
            'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
            'herramienta_especial'
        ]
    )
    
    # Aplicar filtros
    if search_query:
        herramientas = herramientas.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(modelo__icontains=search_query)
        )
    
    if categoria_filtro:
        herramientas = herramientas.filter(categoria_id=categoria_filtro)
    
    if estado_filtro:
        herramientas = herramientas.filter(estado=estado_filtro)
    
    if criticidad_filtro:
        herramientas = herramientas.filter(criticidad=criticidad_filtro)
    
    if calibracion_filtro == 'requiere':
        herramientas = herramientas.filter(requiere_calibracion=True)
    elif calibracion_filtro == 'vencida':
        # Herramientas con calibración vencida
        herramientas = herramientas.filter(requiere_calibracion=True).extra(
            where=["fecha_ultima_calibracion + INTERVAL frecuencia_calibracion DAY < %s"],
            params=[date.today()]
        )
    
    # Calcular métricas para cada herramienta
    for herramienta in herramientas:
        herramienta.valor_total = herramienta.valor_stock_actual()
        herramienta.necesita_reposicion_flag = herramienta.necesita_reposicion()
        herramienta.dias_vencimiento = herramienta.dias_hasta_vencimiento()
        herramienta.necesita_calibracion_check = herramienta.necesita_calibracion_check()
    
    # Estadísticas
    total_herramientas = herramientas.count()
    herramientas_disponibles = herramientas.filter(estado='disponible').count()
    herramientas_mantenimiento = herramientas.filter(estado='mantenimiento').count()
    herramientas_criticas = herramientas.filter(es_herramienta_critica=True).count()
    herramientas_calibracion = herramientas.filter(requiere_calibracion=True).count()
    
    # Valor total del inventario
    valor_total_inventario = sum(h.valor_total for h in herramientas if h.valor_total)
    
    # Alertas
    alertas = []
    
    if herramientas_mantenimiento > 0:
        alertas.append({
            'tipo': 'warning',
            'icono': 'bi-wrench',
            'titulo': f'{herramientas_mantenimiento} Herramienta(s) en Mantenimiento',
            'descripcion': 'Herramientas actualmente en proceso de mantenimiento.'
        })
    
    if herramientas_criticas > 0:
        alertas.append({
            'tipo': 'danger',
            'icono': 'bi-exclamation-triangle',
            'titulo': f'{herramientas_criticas} Herramienta(s) Crítica(s)',
            'descripcion': 'Herramientas marcadas como críticas para la operación.'
        })
    
    # Obtener datos para filtros
    categorias = CategoriaMaterial.objects.filter(
        tipo_categoria__in=[
            'herramientas_manuales', 'herramientas_electricas', 'instrumentos_medicion',
            'herramientas_precision', 'herramientas_corte', 'equipos_laboratorio'
        ],
        activo=True
    ).order_by('nombre')
    
    # Paginación
    paginator = Paginator(herramientas, 12)  # 12 herramientas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'herramientas': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'criticidad_filtro': criticidad_filtro,
        'calibracion_filtro': calibracion_filtro,
        'categorias': categorias,
        'estados': Material.ESTADO_CHOICES,
        'criticidades': Material.CRITICIDAD_CHOICES,
        'total_herramientas': total_herramientas,
        'herramientas_disponibles': herramientas_disponibles,
        'herramientas_mantenimiento': herramientas_mantenimiento,
        'herramientas_criticas': herramientas_criticas,
        'herramientas_calibracion': herramientas_calibracion,
        'valor_total_inventario': valor_total_inventario,
        'alertas': alertas,
        'titulo': 'Herramientas',
    }
    
    return render(request, 'sistema_interno/herramientas.html', context)

@login_required
def crear_material_view(request):
    """Vista para crear nuevo material"""
    
    if request.method == 'POST':
        print(f"DEBUG CREAR: POST recibido - datos: {request.POST}")  # Debug
        form = MaterialForm(request.POST, request.FILES)
        
        print(f"DEBUG CREAR: Form errors: {form.errors}")  # Debug
        print(f"DEBUG CREAR: Form is valid: {form.is_valid()}")  # Debug
        
        if form.is_valid():
            print("DEBUG CREAR: Formulario válido, guardando...")  # Debug
            try:
                material = form.save()
                print(f"DEBUG CREAR: Material guardado: {material.codigo}")  # Debug
                messages.success(
                    request, 
                    f'✅ Material "{material.codigo} - {material.nombre}" creado exitosamente.'
                )
                return redirect('materiales:material-detalle', pk=material.pk)
            except Exception as e:
                print(f"DEBUG CREAR: Error al guardar: {e}")  # Debug
                messages.error(
                    request, 
                    f'Error al guardar el material: {str(e)}'
                )
        else:
            print(f"DEBUG CREAR: Errores del formulario: {form.errors}")  # Debug
            # Mostrar errores específicos
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            
            messages.error(
                request, 
                'Error al crear el material. Por favor revise los campos marcados.'
            )
    else:
        form = MaterialForm()
    
    # Estadísticas para el contexto
    stats = {
        'total_materiales': Material.objects.count(),
        'total_categorias': CategoriaMaterial.objects.filter(activo=True).count(),
        'total_proveedores': Proveedor.objects.filter(activo=True).count(),
    }
    
    context = {
        'form': form,
        'accion': 'crear',
        'titulo': 'Crear Nuevo Material',
        'stats': stats,
    }
    
    return render(request, 'sistema_interno/crear_material.html', context)

@login_required
def crear_herramienta_view(request):
    """Vista para crear nueva herramienta"""
    
    if request.method == 'POST':
        print(f"DEBUG: POST recibido - datos: {request.POST}")  # Debug
        form = HerramientaForm(request.POST, request.FILES)
        
        print(f"DEBUG: Form errors: {form.errors}")  # Debug
        print(f"DEBUG: Form is valid: {form.is_valid()}")  # Debug
        
        if form.is_valid():
            try:
                herramienta = form.save(commit=False)
                
                # Asignar valores por defecto si no se proporcionaron
                if not herramienta.unidad_medida:
                    herramienta.unidad_medida = 'unidad'
                
                # ✅ NO VALIDAR FRECUENCIA_MANTENIMIENTO - simplemente guardar
                # Si requiere mantenimiento pero no se especificó frecuencia, usar un valor por defecto
                if herramienta.requiere_mantenimiento and not herramienta.frecuencia_mantenimiento:
                    herramienta.frecuencia_mantenimiento = 30  # 30 días por defecto
                
                # Si requiere calibración pero no se especificó frecuencia, usar un valor por defecto
                if herramienta.requiere_calibracion and not herramienta.frecuencia_calibracion:
                    herramienta.frecuencia_calibracion = 365  # 1 año por defecto
                
                herramienta.save()
                
                print(f"DEBUG: Herramienta creada exitosamente - {herramienta.codigo}")  # Debug
                
                messages.success(
                    request, 
                    f'La herramienta "{herramienta.nombre}" ha sido creada exitosamente con código {herramienta.codigo}.'
                )
                return redirect('materiales:herramientas')
                
            except Exception as e:
                print(f"DEBUG: Error al guardar herramienta: {e}")  # Debug
                messages.error(request, f'Error al crear la herramienta: {str(e)}')
                
        else:
            print(f"DEBUG: Formulario inválido - {form.errors}")  # Debug
            messages.error(request, 'Error de validación. Por favor revise los campos marcados.')
            
    else:
        form = HerramientaForm()
    
    # Estadísticas para el contexto
    stats = {
        'total_herramientas': Material.objects.filter(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).count(),
        'total_categorias_herramientas': CategoriaMaterial.objects.filter(
            tipo_categoria__in=[
                'herramientas_manuales', 'herramientas_electricas', 'instrumentos_medicion',
                'herramientas_precision', 'herramientas_corte', 'equipos_laboratorio',
                'herramientas_soldadura_eq', 'herramientas_neumaticas', 'herramientas_hidraulicas',
                'instrumentos_calibracion', 'herramientas_seguridad', 'herramientas_especiales'
            ],
            activo=True
        ).count(),
        'total_proveedores': Proveedor.objects.filter(activo=True).count(),
    }
    
    context = {
        'form': form,
        'accion': 'crear',
        'titulo': 'Crear Nueva Herramienta',
        'stats': stats,
    }
    
    return render(request, 'sistema_interno/crear_herramientas.html', context)

@login_required
def material_detalle_view(request, pk):
    """Vista de detalle para material (NO herramienta)"""
    material = get_object_or_404(Material, pk=pk)
    
    # Verificar que efectivamente NO es una herramienta
    if material.es_herramienta():
        # Si es herramienta, redirigir a la vista de detalle de herramientas
        return redirect('materiales:herramienta-detalle', pk=pk)
    
    # Obtener movimientos recientes
    movimientos_recientes = material.movimientos.all().order_by('-fecha_movimiento')[:10]
    
    # Calcular métricas
    material.valor_total = material.valor_stock_actual()
    material.dias_vencimiento = material.dias_hasta_vencimiento()
    material.necesita_reposicion_flag = material.necesita_reposicion()
    
    context = {
        'material': material,
        'movimientos_recientes': movimientos_recientes,
        'es_material': True,  # ✅ FLAG para saber que es material
        'titulo': f'Detalle - {material.nombre}',
    }
    
    return render(request, 'sistema_interno/material_detalle.html', context)

@login_required
def herramienta_detalle_view(request, pk):
    """Vista de detalle específica para herramientas"""
    herramienta = get_object_or_404(Material, pk=pk)
    
    # Verificar que efectivamente ES una herramienta
    if not herramienta.es_herramienta():
        # Si no es herramienta, redirigir a la vista de detalle de materiales
        return redirect('materiales:material-detalle', pk=pk)
    
    # Obtener movimientos recientes
    movimientos_recientes = herramienta.movimientos.all().order_by('-fecha_movimiento')[:10]
    
    # Calcular métricas específicas para herramientas
    herramienta.valor_total = herramienta.valor_stock_actual()
    herramienta.dias_vencimiento = herramienta.dias_hasta_vencimiento()
    herramienta.necesita_reposicion_flag = herramienta.necesita_reposicion()
    herramienta.necesita_calibracion_flag = herramienta.necesita_calibracion_check()
    
    # Días hasta próxima calibración
    if herramienta.requiere_calibracion and herramienta.fecha_ultima_calibracion and herramienta.frecuencia_calibracion:
        proxima_calibracion = herramienta.fecha_ultima_calibracion + timedelta(days=herramienta.frecuencia_calibracion)
        herramienta.dias_proxima_calibracion = (proxima_calibracion - date.today()).days
        # ✅ AGREGAR VALOR ABSOLUTO
        herramienta.dias_proxima_calibracion_abs = abs(herramienta.dias_proxima_calibracion)
    else:
        herramienta.dias_proxima_calibracion = None
        herramienta.dias_proxima_calibracion_abs = None
    
    # Días hasta próximo mantenimiento
    if herramienta.requiere_mantenimiento and herramienta.fecha_ultimo_mantenimiento and herramienta.frecuencia_mantenimiento:
        proximo_mantenimiento = herramienta.fecha_ultimo_mantenimiento + timedelta(days=herramienta.frecuencia_mantenimiento)
        herramienta.dias_proximo_mantenimiento = (proximo_mantenimiento - date.today()).days
        # ✅ AGREGAR VALOR ABSOLUTO
        herramienta.dias_proximo_mantenimiento_abs = abs(herramienta.dias_proximo_mantenimiento)
    else:
        herramienta.dias_proximo_mantenimiento = None
        herramienta.dias_proximo_mantenimiento_abs = None
    
    context = {
        'herramienta': herramienta,
        'material': herramienta,  # Para compatibilidad con templates existentes
        'movimientos_recientes': movimientos_recientes,
        'es_herramienta': True,  # ✅ FLAG para saber que es herramienta
        'titulo': f'Detalle Herramienta - {herramienta.nombre}',
    }
    
    return render(request, 'sistema_interno/herramienta_detalle.html', context)

@login_required
def editar_material_view(request, pk):
    """Vista para editar material existente"""
    
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            material_actualizado = form.save()
            messages.success(
                request, 
                f'✅ Material "{material_actualizado.codigo}" actualizado exitosamente.'
            )
            return redirect('materiales:material-detalle', pk=material_actualizado.pk)
        else:
            messages.error(
                request, 
                'Error al actualizar el material. Por favor revise los campos marcados.'
            )
    else:
        form = MaterialForm(instance=material)
    
    context = {
        'form': form,
        'material': material,
        'accion': 'editar',
        'titulo': f'Editar Material - {material.codigo}',
    }
    
    return render(request, 'sistema_interno/crear_material.html', context)

@login_required
def editar_herramienta_view(request, pk):
    """Vista para editar herramienta existente"""
    
    herramienta = get_object_or_404(Material, pk=pk)
    
    # Verificar que es una herramienta
    if not herramienta.es_herramienta():
        messages.error(request, 'El elemento solicitado no es una herramienta.')
        return redirect('materiales:materiales')
    
    if request.method == 'POST':
        print(f"DEBUG EDITAR HERRAMIENTA: POST recibido - datos: {request.POST}")  # Debug
        form = HerramientaForm(request.POST, request.FILES, instance=herramienta)
        
        print(f"DEBUG EDITAR HERRAMIENTA: Form errors: {form.errors}")  # Debug
        print(f"DEBUG EDITAR HERRAMIENTA: Form is valid: {form.is_valid()}")  # Debug
        
        if form.is_valid():
            print("DEBUG EDITAR HERRAMIENTA: Formulario válido, guardando...")  # Debug
            try:
                herramienta_actualizada = form.save()
                print(f"DEBUG EDITAR HERRAMIENTA: Herramienta actualizada: {herramienta_actualizada.codigo}")  # Debug
                messages.success(
                    request, 
                    f'✅ Herramienta "{herramienta_actualizada.codigo}" actualizada exitosamente.'
                )
                return redirect('materiales:herramienta-detalle', pk=herramienta_actualizada.pk)
            except Exception as e:
                print(f"DEBUG EDITAR HERRAMIENTA: Error al guardar: {e}")  # Debug
                messages.error(
                    request, 
                    f'Error al actualizar la herramienta: {str(e)}'
                )
        else:
            print(f"DEBUG EDITAR HERRAMIENTA: Errores del formulario: {form.errors}")  # Debug
            messages.error(
                request, 
                'Error al actualizar la herramienta. Por favor revise los campos marcados.'
            )
    else:
        form = HerramientaForm(instance=herramienta)
    
    context = {
        'form': form,
        'herramienta': herramienta,
        'material': herramienta,  # Para compatibilidad
        'accion': 'editar',
        'titulo': f'Editar Herramienta - {herramienta.codigo}',
    }
    
    return render(request, 'sistema_interno/crear_herramientas.html', context)

@login_required
def eliminar_material_view(request, pk):
    """Vista para eliminar material"""
    
    material = get_object_or_404(Material, pk=pk)
    
    if request.method == 'POST':
        nombre_material = material.nombre
        codigo_material = material.codigo
        material.delete()
        messages.success(
            request, 
            f'✅ Material "{codigo_material} - {nombre_material}" eliminado exitosamente.'
        )
        return redirect('materiales:materiales')
    
    context = {
        'material': material,
        'titulo': f'Eliminar Material - {material.codigo}',
    }
    
    return render(request, 'sistema_interno/confirmar_eliminar_material.html', context)

@login_required
def eliminar_herramienta_view(request, pk):
    """Vista para eliminar herramienta"""
    
    herramienta = get_object_or_404(Material, pk=pk)
    
    # Verificar que es una herramienta
    if not herramienta.es_herramienta():
        messages.error(request, 'El elemento solicitado no es una herramienta.')
        return redirect('materiales:herramientas')
    
    if request.method == 'POST':
        nombre_herramienta = herramienta.nombre
        codigo_herramienta = herramienta.codigo
        herramienta.delete()
        messages.success(
            request, 
            f'✅ Herramienta "{codigo_herramienta} - {nombre_herramienta}" eliminada exitosamente.'
        )
        return redirect('materiales:herramientas')
    
    context = {
        'herramienta': herramienta,
        'titulo': f'Eliminar Herramienta - {herramienta.codigo}',
    }
    
    return render(request, 'sistema_interno/confirmar_eliminar_herramienta.html', context)

@login_required
def crear_movimiento_view(request, material_pk=None):
    """Vista para crear movimiento de material"""
    
    material = None
    if material_pk:
        material = get_object_or_404(Material, pk=material_pk)
    
    if request.method == 'POST':
        form = MovimientoMaterialForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.usuario = request.user
            movimiento.stock_anterior = movimiento.material.stock_actual
            
            # Calcular nuevo stock según el tipo de movimiento
            if movimiento.tipo_movimiento in ['entrada', 'ajuste_positivo', 'devolucion']:
                nuevo_stock = movimiento.stock_anterior + movimiento.cantidad
            else:
                nuevo_stock = movimiento.stock_anterior - movimiento.cantidad
            
            movimiento.stock_nuevo = max(nuevo_stock, 0)  # No permitir stock negativo
            movimiento.costo_total = movimiento.cantidad * movimiento.costo_unitario
            
            # Guardar movimiento
            movimiento.save()
            
            # Actualizar stock del material
            movimiento.material.stock_actual = movimiento.stock_nuevo
            movimiento.material.save()
            
            messages.success(
                request,
                f'✅ Movimiento registrado exitosamente. Stock actualizado: {movimiento.stock_nuevo}'
            )
            return redirect('materiales:material-detalle', pk=movimiento.material.pk)
        else:
            messages.error(
                request,
                'Error al registrar el movimiento. Por favor revise los campos marcados.'
            )
    else:
        form = MovimientoMaterialForm()
        if material:
            form.fields['material'].initial = material
    
    context = {
        'form': form,
        'material': material,
        'titulo': 'Registrar Movimiento de Material',
    }
    
    return render(request, 'sistema_interno/crear_movimiento.html', context)

@login_required
def movimientos_view(request):
    """Vista para listar movimientos de materiales"""
    
    # Filtros
    material_filtro = request.GET.get('material', '')
    tipo_filtro = request.GET.get('tipo', '')
    motivo_filtro = request.GET.get('motivo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Query base
    movimientos = MovimientoMaterial.objects.select_related('material', 'usuario').all()
    
    # Aplicar filtros
    if material_filtro:
        movimientos = movimientos.filter(material_id=material_filtro)
    
    if tipo_filtro:
        movimientos = movimientos.filter(tipo_movimiento=tipo_filtro)
    
    if motivo_filtro:
        movimientos = movimientos.filter(motivo=motivo_filtro)
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_hasta)
    
    # Ordenar por fecha descendente
    movimientos = movimientos.order_by('-fecha_movimiento')
    
    # Paginación
    paginator = Paginator(movimientos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'movimientos': page_obj,
        'page_obj': page_obj,
        'material_filtro': material_filtro,
        'tipo_filtro': tipo_filtro,
        'motivo_filtro': motivo_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'materiales': Material.objects.filter(activo=True).order_by('nombre'),
        'tipos_movimiento': MovimientoMaterial.TIPO_MOVIMIENTO_CHOICES,
        'motivos': MovimientoMaterial.MOTIVO_CHOICES,
        'titulo': 'Movimientos de Materiales',
    }
    
    return render(request, 'sistema_interno/movimientos.html', context)

@login_required
def exportar_excel_view(request):
    """Vista para exportar materiales a Excel"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Obtener filtros
    tipo = request.GET.get('tipo', 'materiales')  # materiales o herramientas
    
    # Query según el tipo
    if tipo == 'herramientas':
        items = Material.objects.filter(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        filename = f'herramientas_{date.today().strftime("%Y%m%d")}.xlsx'
    else:
        items = Material.objects.exclude(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        filename = f'materiales_{date.today().strftime("%Y%m%d")}.xlsx'
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = [
        'Código', 'Nombre', 'Descripción', 'Tipo', 'Categoría', 'Marca', 'Modelo',
        'Stock Actual', 'Stock Mínimo', 'Precio Unitario', 'Estado', 'Criticidad'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.codigo)
        ws.cell(row=row_num, column=2, value=item.nombre)
        ws.cell(row=row_num, column=3, value=item.descripcion)
        ws.cell(row=row_num, column=4, value=item.get_tipo_display())
        ws.cell(row=row_num, column=5, value=item.categoria.nombre if item.categoria else '')
        ws.cell(row=row_num, column=6, value=item.marca or '')
        ws.cell(row=row_num, column=7, value=item.modelo or '')
        ws.cell(row=row_num, column=8, value=float(item.stock_actual))
        ws.cell(row=row_num, column=9, value=float(item.stock_minimo))
        ws.cell(row=row_num, column=10, value=float(item.precio_unitario))
        ws.cell(row=row_num, column=11, value=item.get_estado_display())
        ws.cell(row=row_num, column=12, value=item.get_criticidad_display())
    
    # Ajustar anchos de columna
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def exportar_pdf_view(request):
    """Vista para exportar materiales a PDF"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from weasyprint import HTML
    
    # Obtener filtros
    tipo = request.GET.get('tipo', 'materiales')
    
    # Query según el tipo
    if tipo == 'herramientas':
        items = Material.objects.filter(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        titulo = 'Inventario de Herramientas'
        filename = f'herramientas_{date.today().strftime("%Y%m%d")}.pdf'
    else:
        items = Material.objects.exclude(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        titulo = 'Inventario de Materiales'
        filename = f'materiales_{date.today().strftime("%Y%m%d")}.pdf'
    
    # Calcular estadísticas
    total_items = items.count()
    valor_total = sum(item.valor_stock_actual() for item in items)
    
    # Renderizar HTML
    html_string = render_to_string('sistema_interno/inventario_pdf_template.html', {
        'items': items,
        'titulo': titulo,
        'total_items': total_items,
        'valor_total': valor_total,
        'fecha_generacion': date.today(),
    })
    
    # Convertir a PDF
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    # Respuesta HTTP
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def exportar_csv_view(request):
    """Vista para exportar materiales a CSV"""
    import csv
    from django.http import HttpResponse
    
    # Obtener filtros
    tipo = request.GET.get('tipo', 'materiales')
    
    # Query según el tipo
    if tipo == 'herramientas':
        items = Material.objects.filter(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        filename = f'herramientas_{date.today().strftime("%Y%m%d")}.csv'
    else:
        items = Material.objects.exclude(
            tipo__in=[
                'herramienta_manual', 'herramienta_electrica', 'herramienta_precision',
                'herramienta_soldadura', 'herramienta_corte', 'herramienta_medicion',
                'herramienta_seguridad', 'instrumento_laboratorio', 'herramienta_neumatica',
                'herramienta_hidraulica', 'equipo_calibracion', 'instrumento_medicion_digital',
                'herramienta_especial'
            ]
        ).select_related('categoria')
        filename = f'materiales_{date.today().strftime("%Y%m%d")}.csv'
    
    # Crear respuesta CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Agregar BOM para UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Headers
    writer.writerow([
        'Código', 'Nombre', 'Descripción', 'Tipo', 'Categoría', 'Marca', 'Modelo',
        'Stock Actual', 'Stock Mínimo', 'Precio Unitario', 'Estado', 'Criticidad'
    ])
    
    # Datos
    for item in items:
        writer.writerow([
            item.codigo,
            item.nombre,
            item.descripcion,
            item.get_tipo_display(),
            item.categoria.nombre if item.categoria else '',
            item.marca or '',
            item.modelo or '',
            float(item.stock_actual),
            float(item.stock_minimo),
            float(item.precio_unitario),
            item.get_estado_display(),
            item.get_criticidad_display(),
        ])
    
    return response

@login_required
def proveedores_view(request):
    """Vista temporal para proveedores - redirigir a inventario"""
    return redirect('inventario:proveedores')