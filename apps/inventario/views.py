from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, F
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import date, timedelta
from io import BytesIO
import json
import csv
from decimal import Decimal
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import User
from .models import Repuesto, CategoriaRepuesto, Proveedor, MovimientoStock
from .forms import RepuestoForm, ProveedorForm

@login_required
def repuestos_view(request):
    """Vista principal de inventario de repuestos"""
    
    # Filtros
    search_query = request.GET.get('search', '')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    criticidad_filtro = request.GET.get('criticidad', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    stock_filtro = request.GET.get('stock', '')  # bajo, agotado, normal
    
    # Query base
    repuestos = Repuesto.objects.select_related('categoria', 'proveedor_principal').all()
    
    # Aplicar filtros
    if search_query:
        repuestos = repuestos.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query) |
            Q(fabricante__icontains=search_query) |
            Q(codigo_fabricante__icontains=search_query) |
            Q(numero_parte__icontains=search_query)
        )
    
    if categoria_filtro:
        repuestos = repuestos.filter(categoria__id=categoria_filtro)
    
    if estado_filtro:
        repuestos = repuestos.filter(estado=estado_filtro)
    
    if criticidad_filtro:
        repuestos = repuestos.filter(criticidad=criticidad_filtro)
    
    if proveedor_filtro:
        repuestos = repuestos.filter(proveedor_principal__id=proveedor_filtro)
    
    if stock_filtro:
        if stock_filtro == 'agotado':
            repuestos = repuestos.filter(stock_actual=0)
        elif stock_filtro == 'bajo':
            repuestos = repuestos.filter(stock_actual__lte=F('stock_minimo'))
        elif stock_filtro == 'normal':
            repuestos = repuestos.filter(stock_actual__gt=F('stock_minimo'))
    
    # Calcular métricas para cada repuesto
    for repuesto in repuestos:
        repuesto.necesita_reorden = repuesto.necesita_reposicion()
        repuesto.valor_total = repuesto.valor_stock_actual()
    
    # Estadísticas generales
    total_repuestos = repuestos.count()
    repuestos_disponibles = repuestos.filter(estado='disponible').count()
    repuestos_agotados = repuestos.filter(estado='agotado').count()
    repuestos_bajo_stock = repuestos.filter(stock_actual__lte=F('stock_minimo')).count()
    repuestos_criticos = repuestos.filter(criticidad='critica').count()
    
    # Valor total del inventario
    valor_total_inventario = sum(r.valor_total for r in repuestos if r.valor_total)
    
    # Obtener datos para filtros
    categorias = CategoriaRepuesto.objects.filter(activo=True).order_by('nombre')
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    # Paginación
    paginator = Paginator(repuestos, 12)  # 12 repuestos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'repuestos': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'criticidad_filtro': criticidad_filtro,
        'proveedor_filtro': proveedor_filtro,
        'stock_filtro': stock_filtro,
        'categorias': categorias,
        'proveedores': proveedores,
        'estados': Repuesto.ESTADO_CHOICES,
        'criticidades': Repuesto.CRITICIDAD_CHOICES,
        'total_repuestos': total_repuestos,
        'repuestos_disponibles': repuestos_disponibles,
        'repuestos_agotados': repuestos_agotados,
        'repuestos_bajo_stock': repuestos_bajo_stock,
        'repuestos_criticos': repuestos_criticos,
        'valor_total_inventario': valor_total_inventario,
        'alertas': [],
        
        'titulo': 'Inventario de Repuestos',
    }
    
    return render(request, 'sistema_interno/repuestos.html', context)

@login_required
def crear_repuesto_view(request):
    """Vista para crear nuevo repuesto"""
    
    if request.method == 'POST':
        form = RepuestoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                repuesto = form.save(commit=False)
                repuesto.creado_por = request.user
                repuesto.save()
                
                # Guardar relaciones many-to-many
                form.save_m2m()
                
                messages.success(
                    request,
                    f'✅ Repuesto "{repuesto.nombre}" creado exitosamente con código {repuesto.codigo}'
                )
                return redirect('inventario:repuestos')
                
            except Exception as e:
                messages.error(request, f'❌ Error al crear el repuesto: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = RepuestoForm()
    
    # Estadísticas para el contexto
    stats = {
        'total_repuestos': Repuesto.objects.count(),
        'total_categorias': CategoriaRepuesto.objects.filter(activo=True).count(),
        'total_proveedores': Proveedor.objects.filter(activo=True).count(),
    }
    
    context = {
        'form': form,
        'accion': 'crear',
        'titulo': 'Crear Nuevo Repuesto',
        'stats': stats,
    }
    
    return render(request, 'sistema_interno/crear_repuesto.html', context)

@login_required
def editar_repuesto_view(request, pk):
    """Vista para editar repuesto existente"""
    
    repuesto = get_object_or_404(Repuesto, pk=pk)
    
    if request.method == 'POST':
        form = RepuestoForm(request.POST, request.FILES, instance=repuesto)
        if form.is_valid():
            try:
                repuesto_actualizado = form.save()
                
                messages.success(
                    request,
                    f'✅ Repuesto "{repuesto_actualizado.nombre}" actualizado exitosamente'
                )
                return redirect('inventario:repuesto-detalle', pk=repuesto_actualizado.pk)
                
            except Exception as e:
                messages.error(request, f'❌ Error al actualizar el repuesto: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = RepuestoForm(instance=repuesto)
    
    context = {
        'form': form,
        'repuesto': repuesto,
        'accion': 'editar',
        'titulo': f'Editar Repuesto - {repuesto.codigo}',
    }
    
    return render(request, 'sistema_interno/crear_repuesto.html', context)

@login_required
def exportar_repuestos_view(request):
    """Vista para exportar repuestos a Excel"""
    
    # Obtener los mismos filtros que la vista principal
    search_query = request.GET.get('search', '')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    criticidad_filtro = request.GET.get('criticidad', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    stock_filtro = request.GET.get('stock', '')
    formato = request.GET.get('formato', 'excel')  # excel o csv
    
    # Query base con los mismos filtros
    repuestos = Repuesto.objects.select_related('categoria', 'proveedor_principal').all()
    
    # Aplicar los mismos filtros
    if search_query:
        repuestos = repuestos.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query)
        )
    
    if categoria_filtro:
        repuestos = repuestos.filter(categoria__id=categoria_filtro)
    
    if estado_filtro:
        repuestos = repuestos.filter(estado=estado_filtro)
    
    if criticidad_filtro:
        repuestos = repuestos.filter(criticidad=criticidad_filtro)
    
    if proveedor_filtro:
        repuestos = repuestos.filter(proveedor_principal__id=proveedor_filtro)
    
    if stock_filtro:
        if stock_filtro == 'agotado':
            repuestos = repuestos.filter(stock_actual=0)
        elif stock_filtro == 'bajo':
            repuestos = repuestos.filter(stock_actual__lte=F('stock_minimo'))
        elif stock_filtro == 'normal':
            repuestos = repuestos.filter(stock_actual__gt=F('stock_minimo'))
    
    if formato == 'csv':
        return exportar_csv(repuestos)
    else:
        return exportar_excel(repuestos)

@login_required
def stock_critico_view(request):
    """Vista para análisis de stock crítico - repuestos con niveles bajos de inventario"""
    
    # Obtener filtros
    categoria_filtro = request.GET.get('categoria', '')
    criticidad_filtro = request.GET.get('criticidad', '')
    estado_filtro = request.GET.get('estado', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    search = request.GET.get('search', '')
    tipo_alerta = request.GET.get('tipo_alerta', '')
    
    # Query base - todos los repuestos activos
    repuestos = Repuesto.objects.filter(activo=True).select_related(
        'categoria', 'proveedor_principal', 'responsable_tecnico'
    )
    
    # === ANÁLISIS DE STOCK CRÍTICO ===
    # Repuestos con stock por debajo del mínimo
    stock_bajo_minimo = repuestos.filter(
        stock_actual__lt=F('stock_minimo')
    ).exclude(stock_minimo=0)
    
    # Repuestos sin stock (agotados)
    repuestos_agotados = repuestos.filter(stock_actual=0)
    
    # Repuestos cerca del punto de reorden
    cerca_reorden = repuestos.filter(
        stock_actual__lte=F('punto_reorden'),
        stock_actual__gt=0
    ).exclude(punto_reorden=0)
    
    # Repuestos críticos con stock bajo
    criticos_stock_bajo = repuestos.filter(
        Q(es_activo_critico=True) | Q(criticidad='CRITICA'),
        stock_actual__lte=F('stock_minimo')
    )
    
    # Repuestos con alta rotación y stock bajo
    alta_rotacion_bajo = repuestos.filter(
        frecuencia_uso__in=['alta', 'muy_alta', 'diaria'],
        stock_actual__lt=F('stock_minimo') + F('stock_seguridad')
    )
    
    # === APLICAR FILTROS ===
    if search:
        repuestos = repuestos.filter(
            Q(codigo__icontains=search) |
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if categoria_filtro:
        repuestos = repuestos.filter(categoria__id=categoria_filtro)
    
    if criticidad_filtro:
        repuestos = repuestos.filter(criticidad=criticidad_filtro)
    
    if estado_filtro:
        repuestos = repuestos.filter(estado=estado_filtro)
    
    if proveedor_filtro:
        repuestos = repuestos.filter(proveedor_principal__id=proveedor_filtro)
    
    # Filtro por tipo de alerta
    if tipo_alerta == 'agotado':
        repuestos = repuestos_agotados
    elif tipo_alerta == 'bajo_minimo':
        repuestos = stock_bajo_minimo
    elif tipo_alerta == 'cerca_reorden':
        repuestos = cerca_reorden
    elif tipo_alerta == 'criticos':
        repuestos = criticos_stock_bajo
    elif tipo_alerta == 'alta_rotacion':
        repuestos = alta_rotacion_bajo
    
    # Calcular métricas para cada repuesto
    repuestos_con_metricas = []
    for repuesto in repuestos:
        # Determinar nivel de alerta
        if repuesto.stock_actual == 0:
            nivel_alerta = 'critico'
        elif repuesto in criticos_stock_bajo:
            nivel_alerta = 'critico'
        elif repuesto.stock_actual <= repuesto.stock_minimo:
            nivel_alerta = 'alto'
        elif repuesto.stock_actual <= repuesto.punto_reorden:
            nivel_alerta = 'medio'
        else:
            nivel_alerta = 'normal'
        
        repuestos_con_metricas.append({
            'repuesto': repuesto,
            'nivel_alerta': nivel_alerta,
            'dias_stock': repuesto.calcular_dias_stock(),
            'valor_riesgo': (repuesto.stock_minimo - repuesto.stock_actual) * repuesto.precio_unitario if repuesto.stock_actual < repuesto.stock_minimo else 0,
        })
    
    # Ordenar por nivel de criticidad
    orden_criticidad = {'critico': 0, 'alto': 1, 'medio': 2, 'normal': 3}
    repuestos_con_metricas.sort(key=lambda x: orden_criticidad.get(x['nivel_alerta'], 4))
    
    # === ESTADÍSTICAS GENERALES ===
    total_repuestos = Repuesto.objects.filter(activo=True).count()
    total_agotados = repuestos_agotados.count()
    total_bajo_minimo = stock_bajo_minimo.count()
    total_cerca_reorden = cerca_reorden.count()
    total_criticos_bajo = criticos_stock_bajo.count()
    
    # Valor total en riesgo
    valor_total_riesgo = sum([
        (r.precio_unitario or 0) * (r.stock_minimo - r.stock_actual) 
        for r in stock_bajo_minimo 
        if r.stock_minimo > r.stock_actual
    ])
    
    # Porcentaje de cumplimiento de stock
    repuestos_con_stock_ok = repuestos.filter(
        stock_actual__gte=F('stock_minimo')
    ).exclude(stock_minimo=0).count()
    
    total_con_minimo = repuestos.exclude(stock_minimo=0).count()
    porcentaje_cumplimiento = (repuestos_con_stock_ok / max(total_con_minimo, 1) * 100)
    
    # === ANÁLISIS POR CATEGORÍA ===
    categorias_criticas = {}
    for categoria in CategoriaRepuesto.objects.filter(activo=True):
        categoria_repuestos = repuestos.filter(categoria=categoria)
        if categoria_repuestos.exists():
            categorias_criticas[categoria.nombre] = {
                'total': categoria_repuestos.count(),
                'agotados': categoria_repuestos.filter(stock_actual=0).count(),
                'bajo_minimo': categoria_repuestos.filter(stock_actual__lt=F('stock_minimo')).count(),
            }
    
    # === RECOMENDACIONES AUTOMÁTICAS ===
    recomendaciones = []
    
    if total_agotados > 0:
        recomendaciones.append({
            'tipo': 'urgente',
            'titulo': 'Repuestos Agotados',
            'descripcion': f'Hay {total_agotados} repuestos sin stock',
            'accion': 'Generar órdenes de compra inmediatas'
        })
    
    if total_criticos_bajo > 0:
        recomendaciones.append({
            'tipo': 'importante',
            'titulo': 'Repuestos Críticos con Stock Bajo',
            'descripcion': f'{total_criticos_bajo} repuestos críticos necesitan reposición',
            'accion': 'Priorizar compras para repuestos críticos'
        })
    
    if porcentaje_cumplimiento < 80:
        recomendaciones.append({
            'tipo': 'mejora',
            'titulo': 'Bajo Cumplimiento de Stock Mínimo',
            'descripcion': f'Solo {porcentaje_cumplimiento:.1f}% de cumplimiento',
            'accion': 'Revisar políticas de stock mínimo'
        })
    
    if valor_total_riesgo > 10000:
        recomendaciones.append({
            'tipo': 'financiero',
            'titulo': 'Alto Valor en Riesgo',
            'descripcion': f'${valor_total_riesgo:.0f} en valor de stock en riesgo',
            'accion': 'Evaluar impacto financiero y tomar medidas'
        })
    
    # === PROVEEDORES PARA FILTROS ===
    proveedores_disponibles = Proveedor.objects.filter(
        activo=True,
        repuestos_principales__in=repuestos
    ).distinct()
    
    categorias_disponibles = CategoriaRepuesto.objects.filter(
        activo=True,
        repuestos__in=repuestos
    ).distinct()
    
    context = {
        'repuestos_criticos': repuestos_con_metricas,
        'recomendaciones': recomendaciones,
        'total_repuestos': total_repuestos,
        'total_agotados': total_agotados,
        'total_bajo_minimo': total_bajo_minimo,
        'total_cerca_reorden': total_cerca_reorden,
        'total_criticos_bajo': total_criticos_bajo,
        'valor_total_riesgo': valor_total_riesgo,
        'porcentaje_cumplimiento': round(porcentaje_cumplimiento, 1),
        'categorias_criticas': categorias_criticas,
        'categoria_filtro': categoria_filtro,
        'criticidad_filtro': criticidad_filtro,
        'estado_filtro': estado_filtro,
        'proveedor_filtro': proveedor_filtro,
        'search': search,
        'tipo_alerta': tipo_alerta,
        'categorias_disponibles': categorias_disponibles,
        'proveedores_disponibles': proveedores_disponibles,
        'criticidades': Repuesto.CRITICIDAD_CHOICES,
        'estados': Repuesto.ESTADO_CHOICES,
        'titulo': 'Stock Crítico',
        'fecha_actualizacion': timezone.now(),
    }
    
    return render(request, 'sistema_interno/stock_critico.html', context)

def exportar_csv(repuestos):
    """Exportar repuestos a CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="inventario_repuestos_{date.today().strftime("%Y%m%d")}.csv"'
    
    # Agregar BOM para UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Escribir encabezados
    headers = [
        'Código', 'Nombre', 'Descripción', 'Categoría', 'Fabricante',
        'Código Fabricante', 'Número de Parte', 'Stock Actual', 'Stock Mínimo',
        'Stock Máximo', 'Punto Reorden', 'Unidad Medida', 'Precio Unitario',
        'Estado', 'Criticidad', 'Proveedor Principal', 'Ubicación Almacén',
        'Fecha Vencimiento', 'Fecha Última Compra', 'Es Consumible',
        'Requiere Refrigeración', 'Observaciones'
    ]
    writer.writerow(headers)
    
    # Escribir datos
    for repuesto in repuestos:
        row = [
            repuesto.codigo,
            repuesto.nombre,
            repuesto.descripcion,
            repuesto.categoria.nombre if repuesto.categoria else '',
            repuesto.fabricante or '',
            repuesto.codigo_fabricante or '',
            repuesto.numero_parte or '',
            float(repuesto.stock_actual),
            float(repuesto.stock_minimo),
            float(repuesto.stock_maximo),
            float(repuesto.punto_reorden),
            repuesto.get_unidad_medida_display(),
            float(repuesto.precio_unitario),
            repuesto.get_estado_display(),
            repuesto.get_criticidad_display(),
            repuesto.proveedor_principal.nombre if repuesto.proveedor_principal else '',
            repuesto.ubicacion_almacen or '',
            repuesto.fecha_vencimiento.strftime('%Y-%m-%d') if repuesto.fecha_vencimiento else '',
            repuesto.fecha_ultima_compra.strftime('%Y-%m-%d') if repuesto.fecha_ultima_compra else '',
            'Sí' if repuesto.es_consumible else 'No',
            'Sí' if repuesto.requiere_refrigeracion else 'No',
            repuesto.observaciones or ''
        ]
        writer.writerow(row)
    
    return response

def exportar_excel(repuestos):
    """Exportar repuestos a Excel con formato"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Repuestos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    headers = [
        'Código', 'Nombre', 'Descripción', 'Categoría', 'Fabricante',
        'Código Fabricante', 'Número de Parte', 'Stock Actual', 'Stock Mínimo',
        'Stock Máximo', 'Punto Reorden', 'Unidad', 'Precio Unitario',
        'Valor Total', 'Estado', 'Criticidad', 'Proveedor', 'Ubicación',
        'Fecha Vencimiento', 'Días Vencimiento', 'Necesita Reposición'
    ]
    
    # Escribir encabezados con estilo
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Escribir datos
    for row_num, repuesto in enumerate(repuestos, 2):
        # Calcular métricas
        valor_total = repuesto.valor_stock_actual()
        dias_vencimiento = repuesto.dias_hasta_vencimiento()
        necesita_reposicion = repuesto.necesita_reposicion()
        
        data = [
            repuesto.codigo,
            repuesto.nombre,
            repuesto.descripcion,
            repuesto.categoria.nombre if repuesto.categoria else '',
            repuesto.fabricante or '',
            repuesto.codigo_fabricante or '',
            repuesto.numero_parte or '',
            float(repuesto.stock_actual),
            float(repuesto.stock_minimo),
            float(repuesto.stock_maximo),
            float(repuesto.punto_reorden),
            repuesto.get_unidad_medida_display(),
            float(repuesto.precio_unitario),
            float(valor_total),
            repuesto.get_estado_display(),
            repuesto.get_criticidad_display(),
            repuesto.proveedor_principal.nombre if repuesto.proveedor_principal else '',
            repuesto.ubicacion_almacen or '',
            repuesto.fecha_vencimiento if repuesto.fecha_vencimiento else '',
            dias_vencimiento if dias_vencimiento is not None else '',
            'Sí' if necesita_reposicion else 'No'
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = border
            if col in [8, 9, 10, 11, 13, 14]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal='right')
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        if col in [3]:  # Descripción
            ws.column_dimensions[column_letter].width = 40
        elif col in [1, 2, 4, 5, 17, 18]:  # Códigos y nombres
            ws.column_dimensions[column_letter].width = 20
        elif col in [6, 7]:  # Códigos fabricante
            ws.column_dimensions[column_letter].width = 15
        else:
            ws.column_dimensions[column_letter].width = 12
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventario_repuestos_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def repuesto_detalle_view(request, pk):
    """Vista detallada de un repuesto"""
    
    repuesto = get_object_or_404(Repuesto, pk=pk)
    
    # Obtener movimientos recientes usando movimientos_stock en lugar de movimientos
    movimientos_recientes = repuesto.movimientos.order_by('-fecha_movimiento')[:10]
    
    # Calcular métricas del repuesto
    repuesto.valor_total = repuesto.stock_actual * repuesto.precio_unitario if repuesto.precio_unitario else 0
    repuesto.necesita_reposicion = repuesto.stock_actual <= repuesto.punto_reorden
    
    # Alertas
    alertas = []
    
    if repuesto.stock_actual <= 0:
        alertas.append({
            'tipo': 'danger',
            'mensaje': 'Repuesto agotado - Stock en cero',
            'icono': 'bi-exclamation-triangle-fill'
        })
    elif repuesto.stock_actual <= repuesto.stock_minimo:
        alertas.append({
            'tipo': 'warning',
            'mensaje': f'Stock por debajo del mínimo ({repuesto.stock_minimo})',
            'icono': 'bi-exclamation-triangle'
        })
    elif repuesto.stock_actual <= repuesto.punto_reorden:
        alertas.append({
            'tipo': 'info',
            'mensaje': f'Stock cerca del punto de reorden ({repuesto.punto_reorden})',
            'icono': 'bi-info-circle'
        })
    
    if repuesto.fecha_vencimiento:
        dias_vencimiento = (repuesto.fecha_vencimiento - date.today()).days
        if dias_vencimiento <= 0:
            alertas.append({
                'tipo': 'danger',
                'mensaje': 'Repuesto vencido',
                'icono': 'bi-calendar-x'
            })
        elif dias_vencimiento <= 30:
            alertas.append({
                'tipo': 'warning',
                'mensaje': f'Vence en {dias_vencimiento} días',
                'icono': 'bi-calendar-event'
            })
    
    context = {
        'repuesto': repuesto,
        'movimientos_recientes': movimientos_recientes,
        'alertas': alertas,
        'titulo': f'Detalle Repuesto - {repuesto.codigo}',
    }
    
    return render(request, 'sistema_interno/repuesto_detalle.html', context)

@login_required
def dashboard_inventario_view(request):
    """Dashboard principal del inventario"""
    
    # Métricas generales
    total_repuestos = Repuesto.objects.filter(activo=True).count()
    repuestos_criticos = Repuesto.objects.filter(es_activo_critico=True, activo=True).count()
    repuestos_agotados = Repuesto.objects.filter(stock_actual=0, activo=True).count()
    
    # Valor total del inventario
    repuestos_con_valor = Repuesto.objects.filter(activo=True, precio_unitario__gt=0)
    valor_total = sum(
        float(r.stock_actual) * float(r.precio_unitario) 
        for r in repuestos_con_valor 
        if r.stock_actual and r.precio_unitario
    )
    
    # Repuestos que necesitan reposición
    necesitan_reposicion = Repuesto.objects.filter(
        stock_actual__lte=F('punto_reorden'), 
        activo=True
    ).count()
    
    # Movimientos recientes
    movimientos_recientes = MovimientoStock.objects.select_related(
        'repuesto', 'usuario'
    ).order_by('-fecha_movimiento')[:5]
    
    # Top 5 repuestos más utilizados (por movimientos de salida)
    repuestos_mas_utilizados = Repuesto.objects.filter(
        movimientos__tipo_movimiento='salida',
        activo=True
    ).annotate(
        total_salidas=Sum('movimientos__cantidad')
    ).order_by('-total_salidas')[:5]
    
    context = {
        'total_repuestos': total_repuestos,
        'repuestos_criticos': repuestos_criticos,
        'repuestos_agotados': repuestos_agotados,
        'valor_total': valor_total,
        'necesitan_reposicion': necesitan_reposicion,
        'movimientos_recientes': movimientos_recientes,
        'repuestos_mas_utilizados': repuestos_mas_utilizados,
        'titulo': 'Dashboard Inventario',
    }
    
    return render(request, 'sistema_interno/dashboard_inventario.html', context)

@login_required
def reportes_inventario_view(request):
    """Vista funcional de reportes de inventario con datos reales"""
    
    # Obtener filtros del request
    periodo_filtro = request.GET.get('periodo', 'ano_actual')
    categoria_filtro = request.GET.get('categoria', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    stock_filtro = request.GET.get('stock', '')
    
    # Consulta base de repuestos
    repuestos_query = Repuesto.objects.all()
    
    if categoria_filtro:
        repuestos_query = repuestos_query.filter(categoria_id=categoria_filtro)
    
    if proveedor_filtro:
        repuestos_query = repuestos_query.filter(proveedor_principal_id=proveedor_filtro)
    
    # Estadísticas generales
    total_repuestos = repuestos_query.count()
    
    # Calcular valor total del inventario
    valor_total = 0
    for repuesto in repuestos_query:
        if repuesto.precio_unitario and repuesto.stock_actual:
            valor_total += float(repuesto.precio_unitario) * float(repuesto.stock_actual)
    
    # Repuestos por estado de stock
    stock_bajo = repuestos_query.filter(stock_actual__lt=F('stock_minimo')).count()
    stock_critico = repuestos_query.filter(Q(stock_actual=0) | Q(stock_actual__lt=F('stock_minimo') / 2)).count()
    stock_optimo = total_repuestos - stock_bajo - stock_critico
    
    # Distribución por categorías
    categorias_data = repuestos_query.values('categoria__nombre').annotate(count=Count('id')).order_by('-count')
    categorias_chart = {
        'labels': [item['categoria__nombre'] or 'Sin categoría' for item in categorias_data],
        'values': [item['count'] for item in categorias_data]
    }
    
    # Datos simulados para gráficos
    movimientos_data = {
        'labels': ['Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'],
        'entradas': [120, 135, 98, 156, 142, 178],
        'salidas': [89, 95, 76, 134, 128, 145]
    }
    
    # Top repuestos por valor
    repuestos_valor = []
    for repuesto in repuestos_query[:10]:
        if repuesto.precio_unitario and repuesto.stock_actual:
            valor = float(repuesto.precio_unitario) * float(repuesto.stock_actual)
            repuestos_valor.append({'nombre': repuesto.nombre, 'valor': valor})
    
    repuestos_valor.sort(key=lambda x: x['valor'], reverse=True)
    repuestos_valor_chart = {
        'labels': [item['nombre'][:20] for item in repuestos_valor[:10]],
        'values': [item['valor'] for item in repuestos_valor[:10]]
    }
    
    stock_chart = {
        'labels': ['Stock Óptimo', 'Stock Bajo', 'Stock Crítico'],
        'values': [stock_optimo, stock_bajo, stock_critico]
    }
    
    # Categorías y proveedores para filtros
    categorias_disponibles = CategoriaRepuesto.objects.filter(activo=True).order_by('nombre')
    proveedores_disponibles = Proveedor.objects.filter(activo=True).order_by('nombre')
    total_proveedores = Proveedor.objects.filter(activo=True).count()
    
    context = {
        'total_repuestos': total_repuestos,
        'valor_total': valor_total,
        'stock_bajo': stock_bajo,
        'stock_critico': stock_critico,
        'stock_optimo': stock_optimo,
        'total_proveedores': total_proveedores,
        'categorias_data': json.dumps(categorias_chart),
        'movimientos_data': json.dumps(movimientos_data),
        'repuestos_valor_data': json.dumps(repuestos_valor_chart),
        'stock_data': json.dumps(stock_chart),
        'categorias_disponibles': categorias_disponibles,
        'proveedores_disponibles': proveedores_disponibles,
        'periodo_filtro': periodo_filtro,
        'categoria_filtro': categoria_filtro,
        'proveedor_filtro': proveedor_filtro,
        'stock_filtro': stock_filtro,
        'reportes_recientes': [],
        'titulo': 'Reportes de Inventario',
    }
    
    return render(request, 'sistema_interno/reportes_inventario.html', context)

@login_required
def exportar_excel_view(request):
    """Vista para exportar datos a Excel"""
    
    tipo_reporte = request.GET.get('tipo', 'stock')
    filtros = {
        'categoria': request.GET.get('categoria', ''),
        'proveedor': request.GET.get('proveedor', ''),
        'stock': request.GET.get('stock', ''),
    }
    
    # Obtener repuestos
    repuestos = Repuesto.objects.select_related('categoria', 'proveedor_principal').all()
    
    # Aplicar filtros
    if filtros.get('categoria'):
        repuestos = repuestos.filter(categoria_id=filtros['categoria'])
    
    if filtros.get('proveedor'):
        repuestos = repuestos.filter(proveedor_principal_id=filtros['proveedor'])
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo_reporte.title()}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    
    # Headers
    headers = ['Código', 'Nombre', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Precio Unitario', 'Valor Total']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    for row_num, repuesto in enumerate(repuestos[:100], 2):  # Limitar a 100
        valor_total = (repuesto.precio_unitario or 0) * (repuesto.stock_actual or 0)
        datos = [
            repuesto.codigo,
            repuesto.nombre,
            repuesto.categoria.nombre if repuesto.categoria else '',
            float(repuesto.stock_actual),
            float(repuesto.stock_minimo),
            float(repuesto.precio_unitario or 0),
            float(valor_total)
        ]
        
        for col, valor in enumerate(datos, 1):
            ws.cell(row=row_num, column=col, value=valor)
    
    # Ajustar columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Inventario_{tipo_reporte}_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def exportar_pdf_view(request):
    """Vista para exportar datos a PDF"""
    
    tipo_reporte = request.GET.get('tipo', 'stock')
    
    # Obtener datos
    repuestos = Repuesto.objects.select_related('categoria', 'proveedor_principal').all()[:50]
    
    # Calcular estadísticas
    total_repuestos = repuestos.count()
    valor_total = sum([(r.precio_unitario or 0) * (r.stock_actual or 0) for r in repuestos])
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_{date.today().strftime("%Y%m%d")}.pdf"'
    
    # Crear documento
    doc = SimpleDocTemplate(response, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30, alignment=1)
    
    # Título
    story.append(Paragraph("REPORTE DE INVENTARIO", title_style))
    story.append(Spacer(1, 12))
    
    # Información del reporte
    info_data = [
        ['Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M')],
        ['Total de repuestos:', str(total_repuestos)],
        ['Valor total:', f'${valor_total:,.2f}'],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 2*inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Tabla de repuestos
    data = [['Código', 'Nombre', 'Stock', 'Precio', 'Valor Total']]
    
    for repuesto in repuestos:
        valor_total_item = 0
        if repuesto.stock_actual and repuesto.precio_unitario:
            valor_total_item = float(repuesto.stock_actual) * float(repuesto.precio_unitario)
        
        data.append([
            repuesto.codigo,
            repuesto.nombre[:30] + '...' if len(repuesto.nombre) > 30 else repuesto.nombre,
            str(repuesto.stock_actual),
            f'${repuesto.precio_unitario or 0:.2f}',
            f'${valor_total_item:.2f}'
        ])
    
    table = Table(data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    return response

@login_required
def movimientos_stock_view(request):
    """Vista para movimientos de stock de repuestos"""
    
    # Filtros
    search = request.GET.get('search', '')
    tipo_filtro = request.GET.get('tipo', '')
    repuesto_filtro = request.GET.get('repuesto', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    usuario_filtro = request.GET.get('usuario', '')
    estado_filtro = request.GET.get('estado', '')
    
    # Query base - CORREGIR EL SELECT_RELATED
    movimientos = MovimientoStock.objects.select_related(
        'repuesto', 
        'repuesto__categoria',  # Relación a través de repuesto
        'repuesto__proveedor_principal',  # CORRECTO: a través de repuesto
        'usuario'
    ).all()
    
    # Aplicar filtros
    if search:
        movimientos = movimientos.filter(
            Q(repuesto__nombre__icontains=search) |
            Q(repuesto__codigo__icontains=search) |
            Q(numero_movimiento__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    if tipo_filtro:
        movimientos = movimientos.filter(tipo_movimiento=tipo_filtro)
    
    if repuesto_filtro:
        movimientos = movimientos.filter(repuesto__id=repuesto_filtro)
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_hasta)
    
    if usuario_filtro:
        movimientos = movimientos.filter(usuario__id=usuario_filtro)
    
    if estado_filtro:
        movimientos = movimientos.filter(estado=estado_filtro)
    
    # Ordenar por fecha (más recientes primero)
    movimientos = movimientos.order_by('-fecha_movimiento')
    
    # === ESTADÍSTICAS ===
    total_movimientos = movimientos.count()
    
    # Movimientos por tipo
    entradas = movimientos.filter(tipo_movimiento__in=['entrada', 'ajuste_positivo', 'transferencia_entrada', 'devolucion']).count()
    salidas = movimientos.filter(tipo_movimiento__in=['salida', 'ajuste_negativo', 'transferencia_salida', 'merma']).count()
    ajustes = movimientos.filter(tipo_movimiento__in=['ajuste_positivo', 'ajuste_negativo']).count()
    transferencias = movimientos.filter(tipo_movimiento__in=['transferencia_entrada', 'transferencia_salida']).count()
    pendientes = movimientos.filter(estado='pendiente').count()
    
    # Valor total de movimientos
    valor_total_movimientos = sum(
        (mov.costo_total or 0) for mov in movimientos
    )
    
    # Repuestos más movidos (top 5)
    from django.db.models import Sum, Count
    repuestos_mas_movidos = movimientos.values(
        'repuesto__nombre', 'repuesto__codigo'
    ).annotate(
        total_movimientos=Count('id'),
        total_cantidad=Sum('cantidad')
    ).order_by('-total_movimientos')[:5]
    
    # Usuarios más activos
    usuarios_activos = movimientos.values(
        'usuario__first_name', 'usuario__last_name', 'usuario__username'
    ).annotate(
        total_movimientos=Count('id')
    ).order_by('-total_movimientos')[:5]
    
    # === DATOS PARA FILTROS ===
    repuestos_disponibles = Repuesto.objects.filter(activo=True).order_by('nombre')
    usuarios_disponibles = User.objects.filter(is_active=True).order_by('first_name')
    tipos_movimiento = MovimientoStock.TIPO_MOVIMIENTO_CHOICES
    estados = MovimientoStock.ESTADO_CHOICES
    
    # === PAGINACIÓN ===
    from django.core.paginator import Paginator
    paginator = Paginator(movimientos, 20)  # 20 movimientos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'movimientos': page_obj,
        'page_obj': page_obj,
        'search': search,
        'tipo_filtro': tipo_filtro,
        'repuesto_filtro': repuesto_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'usuario_filtro': usuario_filtro,
        'estado_filtro': estado_filtro,
        'repuestos_disponibles': repuestos_disponibles,
        'usuarios_disponibles': usuarios_disponibles,
        'tipos_movimiento': tipos_movimiento,
        'estados': estados,
        'total_movimientos': total_movimientos,
        'entradas': entradas,
        'salidas': salidas,
        'ajustes': ajustes,
        'transferencias': transferencias,
        'pendientes': pendientes,
        'valor_total_movimientos': valor_total_movimientos,
        'repuestos_mas_movidos': repuestos_mas_movidos,
        'usuarios_activos': usuarios_activos,
        'titulo': 'Movimientos de Stock',
        'fecha_actualizacion': timezone.now(),
    }
    
    return render(request, 'sistema_interno/movimientos.html', context)

@login_required
def crear_movimiento_view(request):
    """Vista para crear movimiento de stock de repuestos"""
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            repuesto_id = request.POST.get('repuesto')
            tipo_movimiento = request.POST.get('tipo_movimiento')
            motivo = request.POST.get('motivo')
            cantidad = float(request.POST.get('cantidad', 0))
            costo_unitario = float(request.POST.get('costo_unitario', 0))
            observaciones = request.POST.get('observaciones', '')
            motivo_detalle = request.POST.get('motivo_detalle', '')  # Nuevo campo
            ubicacion_destino = request.POST.get('ubicacion_destino', '')  # Nuevo campo
            
            # Validar datos
            if not repuesto_id or not tipo_movimiento or not motivo or cantidad <= 0:
                messages.error(request, 'Por favor complete todos los campos obligatorios.')
                return redirect('inventario:crear-movimiento')
            
            # Obtener el repuesto
            repuesto = get_object_or_404(Repuesto, pk=repuesto_id)
            
            # Calcular stock anterior y nuevo
            stock_anterior = repuesto.stock_actual
            
            if tipo_movimiento in ['entrada', 'ajuste_positivo', 'transferencia_entrada', 'devolucion']:
                stock_nuevo = stock_anterior + Decimal(str(cantidad))
            elif tipo_movimiento in ['salida', 'ajuste_negativo', 'transferencia_salida', 'merma']:
                stock_nuevo = stock_anterior - Decimal(str(cantidad))
                # Validar que no quede stock negativo
                if stock_nuevo < 0:
                    messages.error(request, f'No hay suficiente stock. Stock disponible: {stock_anterior}')
                    return redirect('inventario:crear-movimiento')
            else:
                stock_nuevo = stock_anterior
            
            # Calcular costo total
            costo_total = Decimal(str(cantidad)) * Decimal(str(costo_unitario))
            
            # Crear el movimiento con todos los campos correctos
            movimiento = MovimientoStock.objects.create(
                repuesto=repuesto,
                tipo_movimiento=tipo_movimiento,
                motivo=motivo,
                motivo_detalle=motivo_detalle,  # Campo agregado
                cantidad=Decimal(str(cantidad)),
                stock_anterior=stock_anterior,
                stock_nuevo=stock_nuevo,
                costo_unitario=Decimal(str(costo_unitario)),
                costo_total=costo_total,
                usuario=request.user,
                observaciones=observaciones,
                ubicacion_destino=ubicacion_destino,  # Campo agregado
                fecha_procesamiento=timezone.now(),  # Campo agregado
                estado='procesado'
            )
            
            # Actualizar el stock del repuesto
            repuesto.stock_actual = stock_nuevo
            repuesto.save()
            
            messages.success(request, f'Movimiento {movimiento.numero_movimiento} creado exitosamente.')
            return redirect('inventario:movimientos')
            
        except Exception as e:
            messages.error(request, f'Error al crear el movimiento: {str(e)}')
            return redirect('inventario:crear-movimiento')
    
    # GET request - mostrar formulario
    repuestos = Repuesto.objects.filter(activo=True).order_by('nombre')
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre')
    
    # Tipos de movimiento para el template
    tipos_movimiento = MovimientoStock.TIPO_MOVIMIENTO_CHOICES
    motivos = MovimientoStock.MOTIVO_CHOICES
    
    context = {
        'repuestos': repuestos,
        'proveedores': proveedores,
        'tipos_movimiento': tipos_movimiento,
        'motivos': motivos,
        'titulo': 'Crear Movimiento de Stock',
    }
    
    return render(request, 'sistema_interno/crear_movimiento.html', context)

@login_required
def proveedores_view(request):
    """Vista principal de proveedores"""
    
    # Filtros
    search_query = request.GET.get('search', '')
    estado_filtro = request.GET.get('estado', '')
    categoria_filtro = request.GET.get('categoria', '')
    tipo_filtro = request.GET.get('tipo', '')
    calificacion_filtro = request.GET.get('calificacion', '')
    pais_filtro = request.GET.get('pais', '')
    
    # Query base
    proveedores = Proveedor.objects.all()
    
    # Aplicar filtros
    if search_query:
        proveedores = proveedores.filter(
            Q(codigo__icontains=search_query) |
            Q(nombre__icontains=search_query) |
            Q(nombre_comercial__icontains=search_query) |
            Q(contacto_principal__icontains=search_query) |
            Q(nit__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if estado_filtro:
        proveedores = proveedores.filter(estado=estado_filtro)
    
    if categoria_filtro:
        proveedores = proveedores.filter(categoria=categoria_filtro)
    
    if tipo_filtro:
        proveedores = proveedores.filter(tipo_proveedor=tipo_filtro)
    
    if calificacion_filtro:
        if calificacion_filtro == '4_plus':
            proveedores = proveedores.filter(calificacion__gte=4.0)
        elif calificacion_filtro == '3_to_4':
            proveedores = proveedores.filter(calificacion__gte=3.0, calificacion__lt=4.0)
        elif calificacion_filtro == 'below_3':
            proveedores = proveedores.filter(calificacion__lt=3.0)
    
    if pais_filtro:
        proveedores = proveedores.filter(pais__icontains=pais_filtro)
    
    # Calcular métricas para cada proveedor
    for proveedor in proveedores:
        proveedor.dias_sin_compras = proveedor.dias_sin_comprar()
        proveedor.es_confiable = proveedor.es_proveedor_confiable()
        proveedor.necesita_eval = proveedor.necesita_evaluacion()
        proveedor.tiempo_entrega_info = proveedor.get_tiempo_entrega_categoria()
    
    # Estadísticas generales
    total_proveedores = proveedores.count()
    proveedores_activos = proveedores.filter(estado='activo').count()
    proveedores_evaluacion = proveedores.filter(estado='evaluacion').count()
    proveedores_suspendidos = proveedores.filter(estado='suspendido').count()
    
    # Calificación promedio
    calificacion_promedio = proveedores.filter(
        calificacion__gt=0
    ).aggregate(promedio=Avg('calificacion'))['promedio'] or 0
    
    # Proveedores confiables
    proveedores_confiables = sum(1 for p in proveedores if p.es_confiable)
    
    # Tiempo entrega promedio
    tiempo_entrega_promedio = proveedores.filter(
        tiempo_entrega_promedio__gt=0
    ).aggregate(promedio=Avg('tiempo_entrega_promedio'))['promedio'] or 0
    
    # Valor total comprado
    valor_total_comprado = proveedores.aggregate(
        total=Sum('total_comprado')
    )['total'] or 0
    
    # Alertas
    alertas = []
    
    # Proveedores sin evaluar
    sin_evaluar = sum(1 for p in proveedores if p.necesita_eval)
    if sin_evaluar > 0:
        alertas.append({
            'tipo': 'warning',
            'titulo': f'{sin_evaluar} proveedor(es) necesitan evaluación',
            'descripcion': 'Proveedores que requieren evaluación de desempeño',
            'icono': 'bi-clipboard-check'
        })
    
    # Proveedores con baja calificación
    baja_calificacion = proveedores.filter(calificacion__lt=3.0, calificacion__gt=0).count()
    if baja_calificacion > 0:
        alertas.append({
            'tipo': 'danger',
            'titulo': f'{baja_calificacion} proveedor(es) con calificación baja',
            'descripcion': 'Proveedores con calificación menor a 3.0',
            'icono': 'bi-star'
        })
    
    # Proveedores suspendidos
    if proveedores_suspendidos > 0:
        alertas.append({
            'tipo': 'info',
            'titulo': f'{proveedores_suspendidos} proveedor(es) suspendido(s)',
            'descripcion': 'Proveedores que requieren revisión',
            'icono': 'bi-pause-circle'
        })
    
    # Obtener datos únicos para filtros
    paises_disponibles = list(
        proveedores.values_list('pais', flat=True).distinct().order_by('pais')
    )
    
    # Paginación
    paginator = Paginator(proveedores, 12)  # 12 proveedores por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'proveedores': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'estado_filtro': estado_filtro,
        'categoria_filtro': categoria_filtro,
        'tipo_filtro': tipo_filtro,
        'calificacion_filtro': calificacion_filtro,
        'pais_filtro': pais_filtro,
        
        # Opciones para filtros
        'estados': Proveedor.ESTADO_CHOICES,
        'categorias': Proveedor.CATEGORIA_CHOICES,
        'tipos': Proveedor.TIPO_PROVEEDOR_CHOICES,
        'paises_disponibles': [p for p in paises_disponibles if p],
        
        # Estadísticas
        'total_proveedores': total_proveedores,
        'proveedores_activos': proveedores_activos,
        'proveedores_evaluacion': proveedores_evaluacion,
        'proveedores_suspendidos': proveedores_suspendidos,
        'calificacion_promedio': round(calificacion_promedio, 1),
        'proveedores_confiables': proveedores_confiables,
        'tiempo_entrega_promedio': round(tiempo_entrega_promedio, 1),
        'valor_total_comprado': valor_total_comprado,
        'alertas': alertas,
        
        'titulo': 'Gestión de Proveedores',
    }
    
    return render(request, 'sistema_interno/proveedores.html', context)

@login_required
def crear_proveedor_view(request):
    """Vista para crear nuevo proveedor"""
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                proveedor = form.save(commit=False)
                # El código se genera automáticamente en el modelo
                proveedor.save()
                
                messages.success(
                    request,
                    f'✅ Proveedor "{proveedor.nombre}" creado exitosamente con código {proveedor.codigo}'
                )
                return redirect('inventario:proveedores')
                
            except Exception as e:
                messages.error(request, f'❌ Error al crear el proveedor: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ProveedorForm()
    
    # Estadísticas para el contexto
    stats = {
        'total_proveedores': Proveedor.objects.count(),
        'proveedores_activos': Proveedor.objects.filter(estado='activo').count(),
        'proveedores_evaluacion': Proveedor.objects.filter(estado='evaluacion').count(),
        'promedio_calificacion': Proveedor.objects.filter(
            calificacion__gt=0
        ).aggregate(promedio=Avg('calificacion'))['promedio'] or 0,
    }
    
    context = {
        'form': form,
        'accion': 'crear',
        'titulo': 'Crear Nuevo Proveedor',
        'stats': stats,
    }
    
    return render(request, 'sistema_interno/crear_proveedor.html', context)

@login_required
def detalle_proveedor_view(request, pk):
    """Vista detallada del proveedor"""
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    # Calcular métricas adicionales
    proveedor.dias_sin_compras = proveedor.dias_sin_comprar()
    proveedor.es_confiable = proveedor.es_proveedor_confiable()
    proveedor.necesita_eval = proveedor.necesita_evaluacion()
    proveedor.tiempo_entrega_info = proveedor.get_tiempo_entrega_categoria()
    proveedor.resumen_comercial = proveedor.get_resumen_comercial()
    
    # Obtener repuestos asociados
    repuestos_principales = proveedor.repuestos_principales.filter(activo=True)[:10]
    repuestos_alternativos = proveedor.repuestos_alternativos.filter(activo=True)[:5]
    
    context = {
        'proveedor': proveedor,
        'repuestos_principales': repuestos_principales,
        'repuestos_alternativos': repuestos_alternativos,
    }
    
    return render(request, 'sistema_interno/proveedor_detalle.html', context)

@login_required
def editar_proveedor_view(request, pk):
    """Vista para editar proveedor existente"""
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES, instance=proveedor)
        if form.is_valid():
            try:
                proveedor_actualizado = form.save()
                
                messages.success(
                    request,
                    f'✅ Proveedor "{proveedor_actualizado.nombre}" actualizado exitosamente'
                )
                return redirect('inventario:detalle-proveedor', pk=proveedor_actualizado.pk)
                
            except Exception as e:
                messages.error(request, f'❌ Error al actualizar el proveedor: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ProveedorForm(instance=proveedor)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'accion': 'editar',
        'titulo': f'Editar Proveedor - {proveedor.codigo}',
    }
    
    return render(request, 'sistema_interno/crear_proveedor.html', context)

@login_required
def eliminar_proveedor_view(request, pk):
    """Vista para eliminar proveedor"""
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        nombre = proveedor.nombre
        codigo = proveedor.codigo
        
        # Verificar si tiene repuestos asociados
        repuestos_asociados = proveedor.repuestos_principales.count()
        
        if repuestos_asociados > 0:
            messages.warning(
                request,
                f'⚠️ El proveedor "{nombre}" tiene {repuestos_asociados} repuesto(s) asociado(s). '
                f'Se recomienda cambiar su estado a "Inactivo" en lugar de eliminarlo.'
            )
            return redirect('inventario:detalle-proveedor', pk=pk)
        
        proveedor.delete()
        messages.success(request, f'✅ Proveedor "{nombre}" ({codigo}) eliminado exitosamente.')
        return redirect('inventario:proveedores')
    
    context = {
        'proveedor': proveedor,
    }
    
    return render(request, 'sistema_interno/eliminar_proveedor.html', context)

@login_required
def eliminar_repuesto_view(request, pk):
    """Vista para eliminar repuesto"""
    
    repuesto = get_object_or_404(Repuesto, pk=pk)
    
    if request.method == 'POST':
        nombre = repuesto.nombre
        codigo = repuesto.codigo
        repuesto.delete()
        messages.success(request, f'Repuesto "{codigo} - {nombre}" eliminado exitosamente.')
        return redirect('inventario:repuestos')
    
    context = {
        'repuesto': repuesto,
    }
    
    return render(request, 'sistema_interno/confirmar_eliminar_repuesto.html', context)

@login_required
def alertas_stock_api(request):
    """API para obtener alertas de stock en tiempo real"""
    
    try:
        # Repuestos agotados
        repuestos_agotados = Repuesto.objects.filter(
            activo=True,
            stock_actual=0
        ).values('codigo', 'nombre', 'categoria__nombre')
        
        # Repuestos con stock bajo
        repuestos_bajo_stock = Repuesto.objects.filter(
            activo=True,
            stock_actual__lt=F('stock_minimo'),
            stock_actual__gt=0
        ).values('codigo', 'nombre', 'stock_actual', 'stock_minimo', 'categoria__nombre')
        
        # Repuestos críticos con problemas
        repuestos_criticos = Repuesto.objects.filter(
            activo=True,
            es_activo_critico=True,
            stock_actual__lte=F('stock_minimo')
        ).values('codigo', 'nombre', 'stock_actual', 'stock_minimo')
        
        # Construir respuesta
        alertas = {
            'agotados': {
                'count': repuestos_agotados.count(),
                'items': list(repuestos_agotados[:5])  # Solo primeros 5
            },
            'bajo_stock': {
                'count': repuestos_bajo_stock.count(),
                'items': list(repuestos_bajo_stock[:5])
            },
            'criticos': {
                'count': repuestos_criticos.count(),
                'items': list(repuestos_criticos[:5])
            },
            'total_alertas': repuestos_agotados.count() + repuestos_bajo_stock.count()
        }
        
        return JsonResponse({
            'success': True,
            'alertas': alertas,
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)